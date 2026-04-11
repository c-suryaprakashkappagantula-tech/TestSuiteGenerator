"""
audit_all_features.py — Offline audit of ALL features in DB cache.
Runs every feature through the test engine (no browser needed),
analyzes quality, and produces a comprehensive report.

Usage:  python audit_all_features.py
        python audit_all_features.py --pi PI-52     (single PI only)
        python audit_all_features.py --export       (save report to Excel)
"""
import sys, os, json, re, argparse, time
from datetime import datetime
sys.path.insert(0, os.path.dirname(__file__))

from modules.database import (load_all_features, load_chalk_as_object,
                               get_db_stats, get_chalk_cache_count, _conn)
from modules.test_engine import build_test_suite, _clean_tc_title
from modules.jira_fetcher import JiraIssue
from modules.chalk_parser import ChalkData


def build_mock_jira(feature_id, title=''):
    """Create a minimal JiraIssue for dry-run (no browser needed)."""
    return JiraIssue(
        key=feature_id, summary=title or feature_id,
        description='', status='', priority='', issue_type='Epic',
        assignee='', reporter='', labels=[feature_id],
        pi='', channel='ITMBO', acceptance_criteria='',
    )


def audit_feature(feature_id, pi_label, title):
    """Run a single feature through the engine and return audit results."""
    result = {
        'feature_id': feature_id, 'pi': pi_label, 'title': title[:80],
        'chalk_scenarios': 0, 'chalk_source': 'none',
        'total_tcs': 0, 'total_steps': 0,
        'happy_path': 0, 'negative': 0, 'edge_case': 0, 'e2e': 0,
        'mandatory_negatives': {'hotline': False, 'suspend': False, 'deactivated': False,
                                'invalid_lineid': False, 'invalid_accountid': False, 'mismatch': False},
        'has_transaction_history': False, 'has_century_report': False,
        'has_e2e': False, 'has_test_data_hints': False,
        'tc_name_quality': 'good',  # good / has_long / has_garbage
        'warnings': [],
        'groups': [],
        'status': 'OK',
    }

    # Load Chalk from DB
    chalk = load_chalk_as_object(feature_id, pi_label)
    if not chalk:
        # Try any PI
        c = _conn()
        row = c.execute("SELECT pi_label FROM chalk_cache WHERE feature_id=? AND scenarios_json != '[]' LIMIT 1",
                        (feature_id,)).fetchone()
        c.close()
        if row:
            chalk = load_chalk_as_object(feature_id, row['pi_label'])
            result['chalk_source'] = 'DB (%s)' % row['pi_label']

    if chalk and chalk.scenarios:
        result['chalk_scenarios'] = len(chalk.scenarios)
        result['chalk_source'] = result.get('chalk_source') or 'DB (%s)' % pi_label
    else:
        chalk = ChalkData(feature_id=feature_id)
        result['chalk_source'] = 'none (Jira fallback)'

    # Build test suite
    jira = build_mock_jira(feature_id, title)
    options = {
        'channel': ['ITMBO'], 'devices': ['Mobile'], 'networks': ['4G', '5G'],
        'sim_types': ['eSIM', 'pSIM'], 'os_platforms': ['iOS', 'Android'],
        'include_positive': True, 'include_negative': True,
        'include_e2e': True, 'include_edge': True,
        'include_attachments': False,
        'strategy': 'Smart Suite (Recommended)', 'custom_instructions': '',
    }

    try:
        suite = build_test_suite(jira, chalk, [], options, log=lambda m: None)
    except Exception as e:
        result['status'] = 'ERROR: %s' % str(e)[:60]
        return result

    result['total_tcs'] = len(suite.test_cases)
    result['total_steps'] = sum(len(tc.steps) for tc in suite.test_cases)
    result['warnings'] = suite.warnings

    # Category breakdown
    for tc in suite.test_cases:
        cat = tc.category.lower()
        if 'happy' in cat: result['happy_path'] += 1
        elif 'negative' in cat: result['negative'] += 1
        elif 'edge' in cat: result['edge_case'] += 1
        elif 'e2e' in cat: result['e2e'] += 1

    # Check mandatory negatives
    all_text = ' '.join(tc.summary.lower() + ' ' + tc.description.lower() for tc in suite.test_cases)
    result['mandatory_negatives']['hotline'] = 'hotline' in all_text
    result['mandatory_negatives']['suspend'] = 'suspend' in all_text
    result['mandatory_negatives']['deactivated'] = 'deactiv' in all_text
    result['mandatory_negatives']['invalid_lineid'] = 'lineid' in all_text and 'invalid' in all_text
    result['mandatory_negatives']['invalid_accountid'] = 'accountid' in all_text and 'invalid' in all_text
    result['mandatory_negatives']['mismatch'] = 'mismatch' in all_text

    # Check verification layers
    result['has_transaction_history'] = 'transaction history' in all_text
    result['has_century_report'] = any(kw in all_text for kw in ['century report', 'service grouping'])
    result['has_e2e'] = any(kw in all_text for kw in ['end-to-end', 'e2e'])
    result['has_test_data_hints'] = 'test data:' in all_text.lower()

    # TC name quality
    long_names = sum(1 for tc in suite.test_cases if len(tc.summary) > 120)
    garbage_names = sum(1 for tc in suite.test_cases
                        if 'open item:' in tc.summary.lower() or 'verify: verify' in tc.summary.lower())
    if garbage_names > 0:
        result['tc_name_quality'] = 'has_garbage (%d)' % garbage_names
    elif long_names > 0:
        result['tc_name_quality'] = 'has_long (%d)' % long_names
    else:
        result['tc_name_quality'] = 'good'

    # Groups
    result['groups'] = ['%s:%d' % (g, len(tcs)) for g, tcs in suite.groups.items()]

    return result


def run_audit(pi_filter=None, export=False):
    stats = get_db_stats()
    chalk_count = get_chalk_cache_count()
    feats = load_all_features()

    print('=' * 80)
    print('  TSG FULL AUDIT — All Features from DB Cache')
    print('  DB: %d features | %d Chalk cached | %dKB' % (
        stats['feature_count'], chalk_count, stats['db_size_kb']))
    print('=' * 80)

    # Filter
    if pi_filter:
        feats = {k: v for k, v in feats.items() if k == pi_filter}

    all_results = []
    total = sum(len(fl) for fl in feats.values())
    done = 0
    t0 = time.time()

    for pi, fl in sorted(feats.items()):
        print('\n--- %s (%d features) ---' % (pi, len(fl)))
        for fid, title in fl:
            done += 1
            result = audit_feature(fid, pi, title)
            all_results.append(result)

            # Status indicator
            status_icon = '✅' if result['status'] == 'OK' and result['chalk_scenarios'] > 0 else \
                          '⚠️' if result['status'] == 'OK' and result['chalk_scenarios'] == 0 else '❌'
            mandatory_count = sum(1 for v in result['mandatory_negatives'].values() if v)

            print('  [%d/%d] %s %s: %d chalk → %d TCs (%d steps) | neg:%d mand:%d/6 | %s' % (
                done, total, status_icon, fid,
                result['chalk_scenarios'], result['total_tcs'], result['total_steps'],
                result['negative'], mandatory_count, result['tc_name_quality']))

    elapsed = time.time() - t0

    # ================================================================
    # SUMMARY REPORT
    # ================================================================
    print('\n' + '=' * 80)
    print('  AUDIT SUMMARY')
    print('=' * 80)

    ok = [r for r in all_results if r['status'] == 'OK']
    errors = [r for r in all_results if r['status'] != 'OK']
    with_chalk = [r for r in ok if r['chalk_scenarios'] > 0]
    no_chalk = [r for r in ok if r['chalk_scenarios'] == 0]

    print('\n  Features audited:     %d' % len(all_results))
    print('  ✅ With Chalk data:    %d' % len(with_chalk))
    print('  ⚠️  No Chalk data:     %d' % len(no_chalk))
    print('  ❌ Errors:             %d' % len(errors))

    # TC stats
    total_tcs = sum(r['total_tcs'] for r in ok)
    total_steps = sum(r['total_steps'] for r in ok)
    print('\n  Total TCs generated:  %d' % total_tcs)
    print('  Total Steps:          %d' % total_steps)
    print('  Avg TCs/feature:      %.1f' % (total_tcs / len(ok) if ok else 0))
    print('  Avg Steps/TC:         %.1f' % (total_steps / total_tcs if total_tcs else 0))

    # Category breakdown
    hp = sum(r['happy_path'] for r in ok)
    neg = sum(r['negative'] for r in ok)
    edge = sum(r['edge_case'] for r in ok)
    e2e = sum(r['e2e'] for r in ok)
    print('\n  Category breakdown:')
    print('    Happy Path:  %d (%.0f%%)' % (hp, hp/total_tcs*100 if total_tcs else 0))
    print('    Negative:    %d (%.0f%%)' % (neg, neg/total_tcs*100 if total_tcs else 0))
    print('    Edge Case:   %d (%.0f%%)' % (edge, edge/total_tcs*100 if total_tcs else 0))
    print('    E2E:         %d (%.0f%%)' % (e2e, e2e/total_tcs*100 if total_tcs else 0))

    # Mandatory negative coverage
    print('\n  Mandatory negative coverage (across all features):')
    for check in ['hotline', 'suspend', 'deactivated', 'invalid_lineid', 'invalid_accountid', 'mismatch']:
        covered = sum(1 for r in with_chalk if r['mandatory_negatives'][check])
        print('    %-20s %d/%d (%.0f%%)' % (check, covered, len(with_chalk),
              covered/len(with_chalk)*100 if with_chalk else 0))

    # Verification layers
    print('\n  Verification layers:')
    for layer, key in [('Transaction History', 'has_transaction_history'),
                       ('Century Report', 'has_century_report'),
                       ('E2E', 'has_e2e'), ('Test Data Hints', 'has_test_data_hints')]:
        covered = sum(1 for r in with_chalk if r[key])
        print('    %-20s %d/%d (%.0f%%)' % (layer, covered, len(with_chalk),
              covered/len(with_chalk)*100 if with_chalk else 0))

    # TC name quality
    print('\n  TC name quality:')
    good = sum(1 for r in ok if r['tc_name_quality'] == 'good')
    has_long = sum(1 for r in ok if 'has_long' in r['tc_name_quality'])
    has_garbage = sum(1 for r in ok if 'has_garbage' in r['tc_name_quality'])
    print('    Good:     %d' % good)
    print('    Has long: %d' % has_long)
    print('    Garbage:  %d' % has_garbage)

    # Features with warnings
    warned = [r for r in ok if r['warnings']]
    if warned:
        print('\n  Features with warnings: %d' % len(warned))
        for r in warned[:5]:
            print('    %s: %s' % (r['feature_id'], r['warnings'][0][:60]))

    # No Chalk data features
    if no_chalk:
        print('\n  Features with NO Chalk data (Jira fallback):')
        for r in no_chalk:
            print('    %s @ %s: %s' % (r['feature_id'], r['pi'], r['title'][:50]))

    # Errors
    if errors:
        print('\n  ERRORS:')
        for r in errors:
            print('    %s: %s' % (r['feature_id'], r['status']))

    print('\n  Audit time: %.0fs' % elapsed)

    # Export to Excel
    if export:
        _export_report(all_results)


def _export_report(results):
    """Export audit results to Excel."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from modules.config import OUTPUTS, ts

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Audit Report'

    headers = ['Feature ID', 'PI', 'Title', 'Chalk Scenarios', 'Total TCs', 'Total Steps',
               'Happy Path', 'Negative', 'Edge Case', 'E2E',
               'Hotline', 'Suspend', 'Deactivated', 'Invalid LineId', 'Invalid AccountId', 'Mismatch',
               'Txn History', 'Century Report', 'E2E Flow', 'Test Data',
               'TC Name Quality', 'Status', 'Warnings']

    hf = Font(bold=True, color='FFFFFF')
    hfill = PatternFill(start_color='0B1D39', end_color='0B1D39', fill_type='solid')
    bdr = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))

    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hf; c.fill = hfill; c.border = bdr
    ws.freeze_panes = 'A2'

    for ri, r in enumerate(results, 2):
        vals = [
            r['feature_id'], r['pi'], r['title'], r['chalk_scenarios'],
            r['total_tcs'], r['total_steps'],
            r['happy_path'], r['negative'], r['edge_case'], r['e2e'],
            '✅' if r['mandatory_negatives']['hotline'] else '❌',
            '✅' if r['mandatory_negatives']['suspend'] else '❌',
            '✅' if r['mandatory_negatives']['deactivated'] else '❌',
            '✅' if r['mandatory_negatives']['invalid_lineid'] else '❌',
            '✅' if r['mandatory_negatives']['invalid_accountid'] else '❌',
            '✅' if r['mandatory_negatives']['mismatch'] else '❌',
            '✅' if r['has_transaction_history'] else '❌',
            '✅' if r['has_century_report'] else '❌',
            '✅' if r['has_e2e'] else '❌',
            '✅' if r['has_test_data_hints'] else '❌',
            r['tc_name_quality'], r['status'],
            '; '.join(r['warnings'][:3]) if r['warnings'] else '',
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.border = bdr
            if ci >= 11 and ci <= 20 and v == '❌':
                c.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            elif ci >= 11 and ci <= 20 and v == '✅':
                c.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

    # Auto-width
    for ci in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + ci) if ci <= 26 else 'A'].width = 15

    out = OUTPUTS / ('AUDIT_REPORT_%s.xlsx' % ts())
    wb.save(str(out))
    print('\n  📊 Report exported: %s' % out)


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Audit all features in DB cache')
    parser.add_argument('--pi', type=str, help='Audit only a specific PI')
    parser.add_argument('--export', action='store_true', help='Export report to Excel')
    args = parser.parse_args()
    run_audit(pi_filter=args.pi, export=args.export)
