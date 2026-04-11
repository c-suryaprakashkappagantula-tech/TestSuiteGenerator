"""
Compare Manual Test Suite (MWTGPROV-3942) vs Generated Test Suite (MWTGPROV-4196).
Both are for Change Port-in MDN (CP & CE) feature.
"""
import openpyxl
import re
import json

MANUAL = r'C:\Users\P3314665\Downloads\MWTGPROV-3942_CP & CE_TMO.xlsx'
GENERATED = r'C:\Users\P3314665\Downloads\MWTGPROV-4196_53.2_Adapt_Change_Port-in_MDN_workflow_to_con_20260410_211356.xlsx'

def extract_tcs(path, label):
    wb = openpyxl.load_workbook(path, data_only=True)
    tcs = []
    for sname in wb.sheetnames:
        ws = wb[sname]
        rows = list(ws.iter_rows(min_row=1, values_only=True))
        if not rows:
            continue
        # Find header row
        header_idx = -1
        for i, row in enumerate(rows):
            row_str = ' '.join(str(c or '').lower() for c in row)
            if 's.no' in row_str or 'summary' in row_str or 'test case' in row_str or 'sno' in row_str:
                header_idx = i
                break
        if header_idx < 0:
            # Try row 1 (0-indexed) as header
            header_idx = 0 if len(rows) > 2 else -1
        if header_idx < 0:
            continue

        headers = [str(c or '').strip().lower() for c in rows[header_idx]]
        # Map columns
        col_map = {}
        for ci, h in enumerate(headers):
            if 's.no' in h or h == 'sno' or h == '#' or 'tc' in h and 'no' in h:
                col_map['sno'] = ci
            elif 'summary' in h or 'test case name' in h or 'scenario' in h or 'test scenario' in h:
                col_map['summary'] = ci
            elif 'description' in h or 'desc' in h:
                col_map['description'] = ci
            elif 'precondition' in h or 'pre-condition' in h or 'pre condition' in h:
                col_map['preconditions'] = ci
            elif 'step' in h and ('summary' in h or '#' in h or 'action' in h or 'detail' in h):
                col_map['step_summary'] = ci
            elif 'expected' in h:
                col_map['expected'] = ci
            elif 'step' in h and 'no' in h:
                col_map['step_no'] = ci
            elif h == 'step #' or h == 'step no':
                col_map['step_no'] = ci
            elif 'category' in h or 'type' in h:
                col_map['category'] = ci
            elif 'label' in h:
                col_map['labels'] = ci

        # Extract TCs
        current_tc = None
        for row in rows[header_idx+1:]:
            row_vals = list(row)
            sno_val = str(row_vals[col_map.get('sno', 0)] or '').strip()
            summary_val = str(row_vals[col_map.get('summary', 1)] or '').strip()

            # New TC starts when S.No or Summary is non-empty
            if sno_val and sno_val not in ('None', ''):
                if current_tc:
                    tcs.append(current_tc)
                current_tc = {
                    'sheet': sname,
                    'sno': sno_val,
                    'summary': summary_val,
                    'description': str(row_vals[col_map.get('description', 2)] or '').strip(),
                    'preconditions': str(row_vals[col_map.get('preconditions', 3)] or '').strip(),
                    'steps': [],
                    'category': str(row_vals[col_map.get('category', -1)] or '').strip() if 'category' in col_map else '',
                }
            elif summary_val and not current_tc:
                current_tc = {
                    'sheet': sname,
                    'sno': sno_val or '?',
                    'summary': summary_val,
                    'description': str(row_vals[col_map.get('description', 2)] or '').strip(),
                    'preconditions': str(row_vals[col_map.get('preconditions', 3)] or '').strip(),
                    'steps': [],
                    'category': '',
                }

            # Collect steps
            if current_tc:
                si_sum = col_map.get('step_summary', 5)
                si_exp = col_map.get('expected', 6)
                step_sum = str(row_vals[si_sum] or '').strip() if si_sum < len(row_vals) else ''
                step_exp = str(row_vals[si_exp] or '').strip() if si_exp < len(row_vals) else ''
                if step_sum or step_exp:
                    current_tc['steps'].append({'summary': step_sum, 'expected': step_exp})

        if current_tc:
            tcs.append(current_tc)

    wb.close()
    return tcs


print('='*90)
print('COMPARISON: Manual (MWTGPROV-3942) vs Generated (MWTGPROV-4196)')
print('Both features: Change Port-in MDN (CP & CE)')
print('='*90)

manual_tcs = extract_tcs(MANUAL, 'Manual')
gen_tcs = extract_tcs(GENERATED, 'Generated')

print('\n' + '─'*90)
print('1. OVERVIEW')
print('─'*90)
print(f'  Manual:    {len(manual_tcs)} TCs across sheets')
print(f'  Generated: {len(gen_tcs)} TCs across sheets')

# Show sheets
manual_sheets = set(tc['sheet'] for tc in manual_tcs)
gen_sheets = set(tc['sheet'] for tc in gen_tcs)
print(f'  Manual sheets:    {sorted(manual_sheets)}')
print(f'  Generated sheets: {sorted(gen_sheets)}')

print('\n' + '─'*90)
print('2. TC NAME COMPARISON (side by side)')
print('─'*90)
print(f'\n  {"MANUAL TC NAMES":<70} | {"GENERATED TC NAMES"}')
print(f'  {"─"*70} | {"─"*70}')
max_rows = max(len(manual_tcs), len(gen_tcs))
for i in range(max_rows):
    m_name = manual_tcs[i]['summary'][:68] if i < len(manual_tcs) else ''
    g_name = gen_tcs[i]['summary'][:68] if i < len(gen_tcs) else ''
    m_sno = manual_tcs[i]['sno'] if i < len(manual_tcs) else ''
    g_sno = gen_tcs[i]['sno'] if i < len(gen_tcs) else ''
    print(f'  {m_sno:>3} {m_name:<66} | {g_sno:>3} {g_name}')

print('\n' + '─'*90)
print('3. TC NAME QUALITY ANALYSIS')
print('─'*90)

# Analyze manual TC name patterns
print('\n  MANUAL TC NAME PATTERNS:')
for tc in manual_tcs:
    print(f'    [{tc["sno"]:>3}] {tc["summary"]}')

print('\n  GENERATED TC NAME PATTERNS:')
for tc in gen_tcs:
    print(f'    [{tc["sno"]:>3}] {tc["summary"][:120]}')

# Pattern analysis
print('\n  PATTERN ANALYSIS:')
# Manual pattern detection
m_has_feature_id = sum(1 for tc in manual_tcs if 'MWTGPROV' in tc['summary'])
m_has_tc_prefix = sum(1 for tc in manual_tcs if re.match(r'.*TC\d+', tc['summary']))
m_has_channel = sum(1 for tc in manual_tcs if any(ch in tc['summary'] for ch in ['ITMBO', 'NBOP']))
m_has_device = sum(1 for tc in manual_tcs if any(d in tc['summary'] for d in ['ANDROID', 'iOS', 'Phone', 'Tablet']))
m_has_network = sum(1 for tc in manual_tcs if any(n in tc['summary'] for n in ['4G', '5G']))
m_avg_len = sum(len(tc['summary']) for tc in manual_tcs) / max(len(manual_tcs), 1)

g_has_feature_id = sum(1 for tc in gen_tcs if 'MWTGPROV' in tc['summary'])
g_has_tc_prefix = sum(1 for tc in gen_tcs if re.match(r'.*TC\d+', tc['summary']))
g_has_channel = sum(1 for tc in gen_tcs if any(ch in tc['summary'] for ch in ['ITMBO', 'NBOP']))
g_has_device = sum(1 for tc in gen_tcs if any(d in tc['summary'] for d in ['ANDROID', 'iOS', 'Phone', 'Tablet']))
g_has_network = sum(1 for tc in gen_tcs if any(n in tc['summary'] for n in ['4G', '5G']))
g_avg_len = sum(len(tc['summary']) for tc in gen_tcs) / max(len(gen_tcs), 1)

print(f'    {"Metric":<35} {"Manual":>10} {"Generated":>10}')
print(f'    {"─"*35} {"─"*10} {"─"*10}')
print(f'    {"Has Feature ID (MWTGPROV)":<35} {m_has_feature_id:>10} {g_has_feature_id:>10}')
print(f'    {"Has TC## prefix":<35} {m_has_tc_prefix:>10} {g_has_tc_prefix:>10}')
print(f'    {"Has Channel (ITMBO/NBOP)":<35} {m_has_channel:>10} {g_has_channel:>10}')
print(f'    {"Has Device (ANDROID/iOS)":<35} {m_has_device:>10} {g_has_device:>10}')
print(f'    {"Has Network (4G/5G)":<35} {m_has_network:>10} {g_has_network:>10}')
print(f'    {"Avg TC name length":<35} {m_avg_len:>10.0f} {g_avg_len:>10.0f}')

print('\n' + '─'*90)
print('4. DESCRIPTION COMPARISON')
print('─'*90)
print('\n  MANUAL DESCRIPTIONS:')
for tc in manual_tcs:
    desc = tc['description'][:150] if tc['description'] else '(empty)'
    print(f'    [{tc["sno"]:>3}] {desc}')

print('\n  GENERATED DESCRIPTIONS:')
for tc in gen_tcs:
    desc = tc['description'][:150] if tc['description'] else '(empty)'
    print(f'    [{tc["sno"]:>3}] {desc}')

# Description quality
m_empty_desc = sum(1 for tc in manual_tcs if not tc['description'] or len(tc['description']) < 10)
g_empty_desc = sum(1 for tc in gen_tcs if not tc['description'] or len(tc['description']) < 10)
m_avg_desc = sum(len(tc['description']) for tc in manual_tcs) / max(len(manual_tcs), 1)
g_avg_desc = sum(len(tc['description']) for tc in gen_tcs) / max(len(gen_tcs), 1)

# Check for generic/repetitive descriptions
g_desc_set = set()
g_repetitive = 0
for tc in gen_tcs:
    d = tc['description'][:80].lower()
    if d in g_desc_set:
        g_repetitive += 1
    g_desc_set.add(d)

print(f'\n    {"Metric":<35} {"Manual":>10} {"Generated":>10}')
print(f'    {"─"*35} {"─"*10} {"─"*10}')
print(f'    {"Empty/short descriptions":<35} {m_empty_desc:>10} {g_empty_desc:>10}')
print(f'    {"Avg description length":<35} {m_avg_desc:>10.0f} {g_avg_desc:>10.0f}')
print(f'    {"Repetitive descriptions":<35} {"N/A":>10} {g_repetitive:>10}')

print('\n' + '─'*90)
print('5. STEP QUALITY COMPARISON')
print('─'*90)
m_total_steps = sum(len(tc['steps']) for tc in manual_tcs)
g_total_steps = sum(len(tc['steps']) for tc in gen_tcs)
m_avg_steps = m_total_steps / max(len(manual_tcs), 1)
g_avg_steps = g_total_steps / max(len(gen_tcs), 1)

print(f'    {"Metric":<35} {"Manual":>10} {"Generated":>10}')
print(f'    {"─"*35} {"─"*10} {"─"*10}')
print(f'    {"Total steps":<35} {m_total_steps:>10} {g_total_steps:>10}')
print(f'    {"Avg steps per TC":<35} {m_avg_steps:>10.1f} {g_avg_steps:>10.1f}')

# Show first 3 TCs steps side by side
for i in range(min(3, len(manual_tcs), len(gen_tcs))):
    print(f'\n  TC {i+1} Steps Comparison:')
    print(f'    MANUAL ({manual_tcs[i]["summary"][:50]}):')
    for s in manual_tcs[i]['steps'][:5]:
        print(f'      Step: {s["summary"][:80]}')
        print(f'      Exp:  {s["expected"][:80]}')
    print(f'    GENERATED ({gen_tcs[i]["summary"][:50]}):')
    for s in gen_tcs[i]['steps'][:5]:
        print(f'      Step: {s["summary"][:80]}')
        print(f'      Exp:  {s["expected"][:80]}')

print('\n' + '─'*90)
print('6. SCENARIO COVERAGE COMPARISON')
print('─'*90)

# Extract key themes from manual TCs
manual_themes = set()
for tc in manual_tcs:
    t = (tc['summary'] + ' ' + tc['description']).lower()
    for theme in ['port-in', 'port in', 'change mdn', 'cp', 'ce', 'android', 'ios',
                  '4g', '5g', 'itmbo', 'nbop', 'esim', 'psim', 'phone', 'tablet',
                  'negative', 'hotline', 'suspend', 'deactivat', 'invalid',
                  'rollback', 'e2e', 'century', 'nbop mig', 'transaction history',
                  'downstream', 'timeout', 'concurrent', 'boundary', 'idempoten',
                  'regression', 'cancel', 'genesis', 'mbo', 'syniverse']:
        if theme in t:
            manual_themes.add(theme)

gen_themes = set()
for tc in gen_tcs:
    t = (tc['summary'] + ' ' + tc['description']).lower()
    for theme in ['port-in', 'port in', 'change mdn', 'cp', 'ce', 'android', 'ios',
                  '4g', '5g', 'itmbo', 'nbop', 'esim', 'psim', 'phone', 'tablet',
                  'negative', 'hotline', 'suspend', 'deactivat', 'invalid',
                  'rollback', 'e2e', 'century', 'nbop mig', 'transaction history',
                  'downstream', 'timeout', 'concurrent', 'boundary', 'idempoten',
                  'regression', 'cancel', 'genesis', 'mbo', 'syniverse']:
        if theme in t:
            gen_themes.add(theme)

print(f'  Manual themes:    {sorted(manual_themes)}')
print(f'  Generated themes: {sorted(gen_themes)}')
print(f'  In Manual ONLY:   {sorted(manual_themes - gen_themes)}')
print(f'  In Generated ONLY:{sorted(gen_themes - manual_themes)}')
print(f'  Common:           {sorted(manual_themes & gen_themes)}')

# Category breakdown
print('\n  CATEGORY BREAKDOWN:')
m_cats = {}
for tc in manual_tcs:
    cat = tc.get('category', '') or 'Unknown'
    # Try to detect from name
    tl = tc['summary'].lower()
    if not cat or cat == 'Unknown':
        if any(kw in tl for kw in ['negative', 'invalid', 'fail', 'reject', 'error']):
            cat = 'Negative'
        elif any(kw in tl for kw in ['e2e', 'end-to-end']):
            cat = 'E2E'
        else:
            cat = 'Happy Path'
    m_cats[cat] = m_cats.get(cat, 0) + 1

g_cats = {}
for tc in gen_tcs:
    tl = tc['summary'].lower() + ' ' + tc.get('category', '').lower()
    if any(kw in tl for kw in ['negative', 'invalid', 'fail', 'reject', 'error', 'hotline', 'suspend', 'deactivat', 'mismatch']):
        cat = 'Negative'
    elif any(kw in tl for kw in ['e2e', 'end-to-end']):
        cat = 'E2E'
    elif any(kw in tl for kw in ['edge', 'boundary', 'concurrent', 'idempoten', 'regression']):
        cat = 'Edge Case'
    else:
        cat = 'Happy Path'
    g_cats[cat] = g_cats.get(cat, 0) + 1

all_cats = sorted(set(list(m_cats.keys()) + list(g_cats.keys())))
print(f'    {"Category":<20} {"Manual":>10} {"Generated":>10}')
print(f'    {"─"*20} {"─"*10} {"─"*10}')
for cat in all_cats:
    print(f'    {cat:<20} {m_cats.get(cat, 0):>10} {g_cats.get(cat, 0):>10}')

print('\n' + '─'*90)
print('7. PRECONDITIONS COMPARISON')
print('─'*90)
print('\n  MANUAL PRECONDITIONS (first 3):')
for tc in manual_tcs[:3]:
    pre = tc['preconditions'][:200] if tc['preconditions'] else '(empty)'
    print(f'    [{tc["sno"]}] {pre}')

print('\n  GENERATED PRECONDITIONS (first 3):')
for tc in gen_tcs[:3]:
    pre = tc['preconditions'][:200] if tc['preconditions'] else '(empty)'
    print(f'    [{tc["sno"]}] {pre}')

print('\n' + '='*90)
print('8. FINDINGS & RECOMMENDATIONS')
print('='*90)
print("""
  This section will be filled after analyzing the output above.
  Run this script and review the detailed comparison.
""")
