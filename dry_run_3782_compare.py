"""Dry-run 3782: compare with sample."""
import sys, os, re
sys.path.insert(0, os.path.dirname(__file__))

from modules.test_engine import build_test_suite
from modules.jira_fetcher import JiraIssue
from modules.database import load_chalk_as_object, _conn
from modules.excel_generator import generate_excel
import openpyxl

# Load Chalk
c = _conn()
row = c.execute("SELECT pi_label FROM chalk_cache WHERE feature_id=? AND scenarios_json != '[]' LIMIT 1",
                ('MWTGPROV-3782',)).fetchone()
c.close()
chalk = load_chalk_as_object('MWTGPROV-3782', row['pi_label']) if row else None
print('Chalk: %d scenarios from %s' % (len(chalk.scenarios) if chalk else 0, row['pi_label'] if row else 'none'))

jira = JiraIssue(key='MWTGPROV-3782', summary='Enable Hotline for a subscriber', pi='PI-51', channel='ITMBO',
    labels=['PI-51'], acceptance_criteria='NSL shall enable hotline for active subscriber lines')

options = {'channel': ['ITMBO'], 'devices': ['Mobile'], 'networks': ['4G','5G'],
    'sim_types': ['eSIM','pSIM'], 'os_platforms': ['iOS','Android'],
    'include_positive': True, 'include_negative': True, 'include_e2e': True,
    'include_edge': True, 'include_attachments': False, 'strategy': 'Smart Suite (Recommended)'}

suite = build_test_suite(jira, chalk, [], options, log=print)
out = generate_excel(suite, log=print)

# Show all TCs
print('\n' + '=' * 70)
print('GENERATED TCs')
print('=' * 70)
for sn, gtcs in suite.groups.items():
    print('\n--- %s (%d TCs) ---' % (sn, len(gtcs)))
    for tc in gtcs:
        print('  %s' % tc.summary[:110])
        print('    Desc: %s' % tc.description[:100])

# Load sample
actual = openpyxl.load_workbook(r'..\Downloads\Actual_MWTGPROV-3782.xlsx', data_only=True)
ws = actual['Sheet1']
sample_tcs = []
for r in range(2, ws.max_row+1):
    s = str(ws.cell(r,1).value or '').strip()
    if s and 'MWTGPROV' in s.upper():
        sample_tcs.append(s[:110])

print('\n' + '=' * 70)
print('COMPARISON')
print('=' * 70)
print('Sample TCs: %d' % len(sample_tcs))
print('Our TCs: %d' % len(suite.test_cases))

print('\nSample TC names (first 15):')
for t in sample_tcs[:15]:
    print('  %s' % t)

print('\nOur TC names (first 15):')
for tc in suite.test_cases[:15]:
    print('  %s' % tc.summary[:110])

# Coverage
our_text = ' '.join(tc.summary.lower() + ' ' + tc.description.lower() for tc in suite.test_cases)
checks = {
    'Enable hotline': 'enable hotline',
    'Hotline removed on deactivate': 'removed when deactivat',
    'Hotline removed on suspend': 'removed when suspend',
    'Not active rejected': 'not active|suspend.*status',
    'Account not exist': 'account.*exist|invalid.*account',
    'Line not exist': 'line.*exist|invalid.*line',
    'MDN not exist': 'mdn.*exist|invalid mdn',
    'Already hotlined': 'already.*hotline',
    'Schema validation': 'schema',
    'ICCID check': 'iccid',
    'Restore suspend fails': 'restore suspend',
    'Outbound calls redirected': 'outbound call',
    'Suspended MDN': 'suspend',
    'Deactivated MDN': 'deactiv',
    'Invalid LineId': 'invalid.*lineid|non-existent.*lineid',
    'Invalid AccountId': 'invalid.*accountid|non-existent.*accountid',
    'Mismatch': 'mismatch',
    'E2E': 'end-to-end|e2e',
    'Transaction History': 'transaction history',
    'Century Report': 'century report|service grouping',
}
covered = 0
for name, pattern in checks.items():
    found = bool(re.search(pattern, our_text))
    if found: covered += 1
    print('  %-35s %s' % (name, 'YES' if found else 'MISSING'))
print('\nCoverage: %d/%d (%.0f%%)' % (covered, len(checks), covered/len(checks)*100))
print('\nOutput: %s' % out)
