"""Compare May 7 (old baseline) vs today's latest for 4166."""
import sys, os, glob
sys.path.insert(0, os.path.dirname(__file__))
import openpyxl

output_dir = os.path.join(os.path.dirname(__file__), 'outputs')

# Specific files
old_candidates = glob.glob(os.path.join(output_dir, "*4166*20260507_214702*"))
if not old_candidates:
    # Try without glob — list dir and filter
    all_files = os.listdir(output_dir)
    old_candidates = [os.path.join(output_dir, f) for f in all_files 
                      if '4166' in f and '20260507_214702' in f and f.endswith('.xlsx')]
old_file = old_candidates[0]

new_candidates = sorted([os.path.join(output_dir, f) for f in os.listdir(output_dir)
                         if '4166' in f and '20260521' in f and f.endswith('.xlsx') and 'CHECKPOINT' not in f])
new_file = new_candidates[-1]

print(f"OLD (May 7):  {os.path.basename(old_file)}")
print(f"NEW (Today):  {os.path.basename(new_file)}")


def extract_tcs(filepath):
    """Extract TCs with steps from Excel."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    tcs = []
    for sn in wb.sheetnames:
        if sn in ('Summary', 'Traceability', 'Combinations'):
            continue
        ws = wb[sn]
        for r in range(3, ws.max_row + 1):
            sno = str(ws.cell(r, 1).value or '').strip()
            summary = str(ws.cell(r, 2).value or '').strip()
            if not summary:
                continue
            precond = str(ws.cell(r, 3).value or '').strip()
            category = str(ws.cell(r, 4).value or '').strip()
            steps = str(ws.cell(r, 5).value or '').strip()
            expected = str(ws.cell(r, 6).value or '').strip()
            tcs.append({
                'sno': sno,
                'summary': summary,
                'preconditions': precond,
                'category': category,
                'steps': steps,
                'expected': expected,
            })
    wb.close()
    return tcs


old_tcs = extract_tcs(old_file)
new_tcs = extract_tcs(new_file)

print(f"\nOLD: {len(old_tcs)} TCs")
print(f"NEW: {len(new_tcs)} TCs")

# Show OLD TCs
print(f"\n{'═'*80}")
print("  OLD SUITE (May 7) — TC summaries + Step 1")
print(f"{'═'*80}")
for i, tc in enumerate(old_tcs, 1):
    print(f"  {i:2d}. [{tc['category'][:12]:12s}] {tc['summary'][:70]}")
    if tc['steps']:
        lines = [l.strip() for l in tc['steps'].replace('\r\n', '\n').split('\n') if l.strip()]
        for l in lines[:2]:
            print(f"       S: {l[:75]}")

# Show NEW TCs
print(f"\n{'═'*80}")
print("  NEW SUITE (Today) — TC summaries + Step 1")
print(f"{'═'*80}")
for i, tc in enumerate(new_tcs, 1):
    print(f"  {i:2d}. [{tc['category'][:12]:12s}] {tc['summary'][:70]}")
    if tc['steps']:
        lines = [l.strip() for l in tc['steps'].replace('\r\n', '\n').split('\n') if l.strip()]
        for l in lines[:2]:
            print(f"       S: {l[:75]}")

# Detailed diff for TC01
print(f"\n{'═'*80}")
print("  DETAILED DIFF — TC01 (CR Fix)")
print(f"{'═'*80}")
if old_tcs and new_tcs:
    # Find the CR fix TC in both
    old_tc1 = old_tcs[0]
    new_tc1 = new_tcs[0]

    print(f"\n  OLD TC01: {old_tc1['summary'][:70]}")
    print(f"  NEW TC01: {new_tc1['summary'][:70]}")

    print(f"\n  OLD Preconditions:")
    for l in old_tc1['preconditions'].replace('\r\n', '\n').split('\n'):
        if l.strip():
            print(f"    {l.strip()[:80]}")

    print(f"\n  NEW Preconditions:")
    for l in new_tc1['preconditions'].replace('\r\n', '\n').split('\n'):
        if l.strip():
            print(f"    {l.strip()[:80]}")

    print(f"\n  OLD Steps:")
    for l in old_tc1['steps'].replace('\r\n', '\n').split('\n'):
        if l.strip():
            print(f"    {l.strip()[:80]}")

    print(f"\n  NEW Steps:")
    for l in new_tc1['steps'].replace('\r\n', '\n').split('\n'):
        if l.strip():
            print(f"    {l.strip()[:80]}")

    print(f"\n  OLD Expected:")
    for l in old_tc1['expected'].replace('\r\n', '\n').split('\n'):
        if l.strip():
            print(f"    {l.strip()[:80]}")

    print(f"\n  NEW Expected:")
    for l in new_tc1['expected'].replace('\r\n', '\n').split('\n'):
        if l.strip():
            print(f"    {l.strip()[:80]}")
