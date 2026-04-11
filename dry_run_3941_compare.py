"""
Dry-run: Build test suite for MWTGPROV-3941 (Port-Out) using mock data,
then compare with the sample file.
"""
import sys, os
sys.path.insert(0, os.path.dirname(__file__))

from modules.config import OUTPUTS
from modules.test_engine import build_test_suite, TestSuite
from modules.jira_fetcher import JiraIssue, JiraAttachment
from modules.chalk_parser import ChalkData, ChalkScenario
from modules.excel_generator import generate_excel

# ================================================================
# MOCK JIRA DATA (based on real 3941 structure)
# ================================================================
jira = JiraIssue(
    key='MWTGPROV-3941',
    summary='Port-Out - Unsolicited Port Out / Update Port Out / Cancel Port Out',
    description="""Port-Out feature for TMO integration.
1. NSL shall process Unsolicited Port-Out (UP) requests from TMO via NE
2. NSL shall forward UP notification to IT-MBO
3. NSL shall process Update Port-Out (PO) requests from MBO
4. NSL shall handle Cancel Port-Out requests
5. NSL shall update Transaction History and MNP Details tables
6. NSL shall validate MDN status before processing port-out
7. NSL shall handle rejection responses from IT-MBO
8. NSL shall trigger port-out deactivation after successful PO
9. NSL shall update NBOP MIG tables after port-out completion
10. NSL shall verify Century Report for all backend calls""",
    status='In Progress',
    priority='High',
    issue_type='Epic',
    assignee='QA Team',
    reporter='Dev Lead',
    labels=['MWTGPROV-3941', 'PI-51', 'PortOut', 'TMO'],
    pi='PI-51',
    channel='ITMBO',
    acceptance_criteria="""
* NSL shall process Unsolicited Port-Out (UP) from TMO via NE and forward to IT-MBO
* NSL shall process Update Port-Out (PO) from MBO and trigger deactivation
* NSL shall handle Cancel Port-Out and update MNP status to Cancelled
* NSL shall reject port-out for Hotlined, Suspended, and Deactivated MDNs
* NSL shall validate MDN exists and is in Active status
* NSL shall handle IT-MBO rejection responses (Invalid PIN 6A, Invalid Account 8A)
* NSL shall update Transaction History: TRANSACTION_STATUS, NOTIFICATION_STATUS, MNP_STATUS
* NSL shall update MNP_DETAILS table with MNP_ID, MDN, MNP_TYPE, MNP_STATUS
* NSL shall trigger port-out deactivation within 1 minute of successful PO
* NSL shall verify all backend calls in Century Report
* NSL shall handle notification timeout (PO not triggered within 50 secs)
* NSL shall support secondary environment routing if MDN not found in primary
""",
)

# ================================================================
# MOCK CHALK DATA (based on real 3941 Chalk page structure)
# ================================================================
chalk = ChalkData(
    feature_id='MWTGPROV-3941',
    feature_title='MWTGPROV-3941 - Port-Out - Unsolicited Port Out / Update Port Out / Cancel Port Out',
    scope="""Port-Out feature handles three transaction types:
UP (Unsolicited Port-Out): TMO sends port-out notification via NE to NSL, NSL forwards to IT-MBO
PO (Update Port-Out): MBO responds with port-out confirmation, NSL triggers deactivation to TMO
Cancel Port-Out: Cancels an in-progress port-out, updates MNP status to Cancelled""",
    rules="""Rules:
1. MDN must be in Active status for port-out processing
2. Hotlined, Suspended, Deactivated MDNs shall be rejected
3. PO must be received within 50 seconds of UP notification
4. Port-out deactivation must be triggered within 1 minute of successful PO
5. Transaction History must be updated at each stage""",
    scenarios=[
        ChalkScenario(
            scenario_id='TS_MWTGPROV-3941_1',
            title='Unsolicited Port-Out (UP) - Happy Path - pSIM - UP then PO then Deactivation',
            prereq='Pre-req: Active TMO subscriber line with pSIM. ITMBO accessible.',
            steps=[
                'Step 1: Trigger Unsolicited PortOut API with trans type "UP"',
                'Step 2: Verify UP call reaches from TMO to NE to NSL to MBO',
                'Step 3: Validate Century Report for UP inbound and outbound calls',
                'Step 4: Validate NBOP_MIG_TRANSACTION_HISTORY: TRANSACTION_STATUS=IN PROGRESS, NOTIFICATION_STATUS=PENDING',
                'Step 5: Trigger Update PortOut API with trans type "PO" within 50 secs',
                'Step 6: Validate MBO responds with Response Type "C" and code 7A',
                'Step 7: Validate NBOP_MIG_TRANSACTION_HISTORY: TRANSACTION_STATUS=COMPLETED, NOTIFICATION_STATUS=SUCCESS',
                'Step 8: Trigger PortOut Deactivation within 1 min',
                'Step 9: Validate deactivation reaches NE then NSL then MBO',
                'Step 10: Validate Century Report for all backend calls',
                'Step 11: Validate NBOP tables: MNP_DETAILS, NBOP_MIG_ACCOUNT, DEVICE, FEATURE, LINE, LINE_HIST',
                'Step 12: Validate TMO-Genesis portal: MDN in deactivated state',
            ],
            validation='Port-out completed successfully. MDN deactivated. All NBOP tables updated. Century Report shows all calls.',
            category='Happy Path',
        ),
        ChalkScenario(
            scenario_id='TS_MWTGPROV-3941_2',
            title='Unsolicited Port-Out (UP) - Happy Path - eSIM - UP then PO then Deactivation',
            prereq='Pre-req: Active TMO subscriber line with eSIM. ITMBO accessible.',
            steps=[
                'Step 1: Trigger Unsolicited PortOut API with trans type "UP"',
                'Step 2: Verify UP call reaches from TMO to NE to NSL to MBO',
                'Step 3: Validate Century Report for UP inbound and outbound calls',
                'Step 4: Trigger Update PortOut API with trans type "PO" within 50 secs',
                'Step 5: Validate MBO responds with Response Type "C"',
                'Step 6: Trigger PortOut Deactivation within 1 min',
                'Step 7: Validate all NBOP tables and Century Report',
                'Step 8: Validate TMO-Genesis portal: MDN in deactivated state',
            ],
            validation='Port-out completed for eSIM. MDN deactivated. All tables updated.',
            category='Happy Path',
        ),
        ChalkScenario(
            scenario_id='TS_MWTGPROV-3941_3',
            title='Cancel Port-Out - Happy Path - Cancel after UP before PO',
            prereq='Pre-req: Active TMO subscriber line. UP already triggered.',
            steps=[
                'Step 1: Trigger Unsolicited PortOut API with trans type "UP"',
                'Step 2: Trigger Cancel Port-Out before PO is received',
                'Step 3: Validate MNP_STATUS updated to Cancelled',
                'Step 4: Validate TRANSACTION_STATUS=COMPLETED, NOTIFICATION_STATUS=SUCCESS',
                'Step 5: Validate Century Report shows cancel transaction',
            ],
            validation='Cancel port-out processed. MNP_STATUS=Cancelled. Transaction completed.',
            category='Happy Path',
        ),
        ChalkScenario(
            scenario_id='TS_MWTGPROV-3941_4',
            title='Cancel Port-Out - Notification FAILED - MNP Status Failure',
            prereq='Pre-req: Active TMO subscriber line. UP already triggered.',
            steps=[
                'Step 1: Trigger Unsolicited PortOut API with trans type "UP"',
                'Step 2: Trigger Cancel Port-Out with notification failure condition',
                'Step 3: Validate NOTIFICATION_STATUS=FAILED, MNP_STATUS=Failure',
            ],
            validation='Cancel port-out notification failed. MNP_STATUS reflects failure.',
            category='Negative',
        ),
        ChalkScenario(
            scenario_id='TS_MWTGPROV-3941_5',
            title='Port-Out - UP then PO Failure - PO response error',
            prereq='Pre-req: Active TMO subscriber line.',
            steps=[
                'Step 1: Trigger Unsolicited PortOut API with trans type "UP"',
                'Step 2: Trigger Update PortOut with error condition from MBO',
                'Step 3: Validate TRANSACTION_STATUS=FAILED',
                'Step 4: Validate Century Report shows failure',
            ],
            validation='PO failure handled. Transaction marked failed.',
            category='Negative',
        ),
        ChalkScenario(
            scenario_id='TS_MWTGPROV-3941_6',
            title='Port-Out - Reject Hotlined MDN',
            prereq='Pre-req: TMO subscriber line in Hotlined status.',
            steps=[
                'Step 1: Trigger Unsolicited PortOut API with Hotlined MDN',
                'Step 2: Validate NSL rejects the port-out request',
                'Step 3: Validate Century Report shows rejection',
            ],
            validation='NSL rejects port-out for Hotlined MDN.',
            category='Negative',
        ),
        ChalkScenario(
            scenario_id='TS_MWTGPROV-3941_7',
            title='Port-Out - Reject Suspended MDN',
            prereq='Pre-req: TMO subscriber line in Suspended status.',
            steps=[
                'Step 1: Trigger Unsolicited PortOut API with Suspended MDN',
                'Step 2: Validate NSL rejects the port-out request',
                'Step 3: Validate Century Report shows rejection',
            ],
            validation='NSL rejects port-out for Suspended MDN.',
            category='Negative',
        ),
        ChalkScenario(
            scenario_id='TS_MWTGPROV-3941_8',
            title='Port-Out - Reject Deactivated MDN',
            prereq='Pre-req: TMO subscriber line in Deactivated status.',
            steps=[
                'Step 1: Trigger Unsolicited PortOut API with Deactivated MDN',
                'Step 2: Validate NSL rejects the port-out request',
            ],
            validation='NSL rejects port-out for Deactivated MDN.',
            category='Negative',
        ),
        ChalkScenario(
            scenario_id='TS_MWTGPROV-3941_9',
            title='Port-Out - MDN Not Exists - ERR07',
            prereq='Pre-req: MDN not registered in TMO network.',
            steps=[
                'Step 1: Trigger Unsolicited PortOut API with non-existent MDN',
                'Step 2: Validate NSL returns ERR07 - MDN Not Found',
                'Step 3: Validate Transaction History: Failure, Notification: Failed',
            ],
            validation='ERR07 - MDN Not Found. Transaction and Notification status: Failed.',
            category='Negative',
        ),
        ChalkScenario(
            scenario_id='TS_MWTGPROV-3941_10',
            title='Unsolicited Port-Out - MDN less than 10 digits',
            prereq='Pre-req: System in ready state.',
            steps=[
                'Step 1: Trigger Unsolicited PortOut API with 9-digit MDN',
                'Step 2: Validate NSL throws error for invalid MDN length',
            ],
            validation='NSL throws error for MDN less than 10 digits.',
            category='Negative',
        ),
        ChalkScenario(
            scenario_id='TS_MWTGPROV-3941_11',
            title='ITMBO rejects UP request - 6A Invalid PIN',
            prereq='Pre-req: Active TMO subscriber line.',
            steps=[
                'Step 1: Trigger Unsolicited PortOut API with trans type "UP"',
                'Step 2: NSL forwards UP to IT-MBO',
                'Step 3: IT-MBO rejects with response code 6A (Invalid PIN)',
                'Step 4: Validate NSL handles rejection and updates transaction status',
            ],
            validation='IT-MBO rejection 6A handled. Transaction status updated to Failed.',
            category='Negative',
        ),
        ChalkScenario(
            scenario_id='TS_MWTGPROV-3941_12',
            title='ITMBO rejects UP request - 8A Invalid Account Number',
            prereq='Pre-req: Active TMO subscriber line.',
            steps=[
                'Step 1: Trigger Unsolicited PortOut API with trans type "UP"',
                'Step 2: NSL forwards UP to IT-MBO',
                'Step 3: IT-MBO rejects with response code 8A (Invalid Account Number)',
                'Step 4: Validate NSL handles rejection and updates transaction status',
            ],
            validation='IT-MBO rejection 8A handled. Transaction status updated to Failed.',
            category='Negative',
        ),
        ChalkScenario(
            scenario_id='TS_MWTGPROV-3941_13',
            title='Port-Out - Notification timeout - PO not triggered within 50 secs',
            prereq='Pre-req: Active TMO subscriber line.',
            steps=[
                'Step 1: Trigger Unsolicited PortOut API with trans type "UP"',
                'Step 2: Wait beyond 50 seconds without triggering PO',
                'Step 3: Validate NOTIFICATION_STATUS=FAILED',
                'Step 4: Validate Transaction History reflects timeout',
            ],
            validation='PO not received within 50 secs. Notification status: FAILED.',
            category='Negative',
        ),
        ChalkScenario(
            scenario_id='TS_MWTGPROV-3941_14',
            title='Unsolicited Port-Out - without transactionType - throw error',
            prereq='Pre-req: System in ready state.',
            steps=[
                'Step 1: Trigger Unsolicited PortOut API without transactionType field',
                'Step 2: Validate NSL throws error',
            ],
            validation='NSL throws error for missing transactionType.',
            category='Negative',
        ),
        ChalkScenario(
            scenario_id='TS_MWTGPROV-3941_15',
            title='Unsolicited Port-Out - without transactionTimeStamp, requestDate, desiredDueDate, lnp - throw error',
            prereq='Pre-req: System in ready state.',
            steps=[
                'Step 1: Trigger Unsolicited PortOut API without transactionTimeStamp, requestDate, desiredDueDate, lnp',
                'Step 2: Validate NSL throws error',
            ],
            validation='NSL throws error for missing required fields.',
            category='Negative',
        ),
        ChalkScenario(
            scenario_id='TS_MWTGPROV-3941_16',
            title='Port-Out - Inquiry Status Rescheduled - line not deactivated in TMO',
            prereq='Pre-req: Active TMO subscriber line. Port-out in progress.',
            steps=[
                'Step 1: Trigger Unsolicited PortOut API',
                'Step 2: Complete UP and PO flow',
                'Step 3: Line not deactivated in TMO within expected time',
                'Step 4: Validate NSL updates ENQUIRY_STATUS as "Rescheduled"',
            ],
            validation='ENQUIRY_STATUS updated to Rescheduled when line not deactivated.',
            category='Edge Case',
        ),
        ChalkScenario(
            scenario_id='TS_MWTGPROV-3941_17',
            title='Port-Out - MDN matches Secondary Environment - E2E from ITMBO',
            prereq='Pre-req: MDN available in SIT (secondary) environment but not primary.',
            steps=[
                'Step 1: Trigger Unsolicited PortOut API from Proxy URL with MDN in SIT env',
                'Step 2: Validate API routes to secondary environment',
                'Step 3: Validate port-out completes in secondary environment',
                'Step 4: Validate Century Report and NBOP tables',
            ],
            validation='Port-out routed to secondary environment. E2E completed successfully.',
            category='E2E',
        ),
        ChalkScenario(
            scenario_id='TS_MWTGPROV-3941_18',
            title='Update PortOut - MDN less than 10 digits',
            prereq='Pre-req: System in ready state.',
            steps=[
                'Step 1: Trigger Update PortOut API with 9-digit MDN',
                'Step 2: Validate MBO throws error for invalid MDN length',
            ],
            validation='MBO throws error for MDN less than 10 digits.',
            category='Negative',
        ),
        ChalkScenario(
            scenario_id='TS_MWTGPROV-3941_19',
            title='Update PortOut - Invalid Line ID',
            prereq='Pre-req: System in ready state.',
            steps=[
                'Step 1: Trigger Update PortOut API with invalid line ID',
                'Step 2: Validate MBO throws error',
            ],
            validation='MBO throws error for invalid line ID.',
            category='Negative',
        ),
        ChalkScenario(
            scenario_id='TS_MWTGPROV-3941_20',
            title='Update PortOut - Invalid Account ID',
            prereq='Pre-req: System in ready state.',
            steps=[
                'Step 1: Trigger Update PortOut API with invalid account ID',
                'Step 2: Validate MBO throws error',
            ],
            validation='MBO throws error for invalid account ID.',
            category='Negative',
        ),
        ChalkScenario(
            scenario_id='TS_MWTGPROV-3941_21',
            title='Update PortOut - LineId and AccountId Mismatch',
            prereq='Pre-req: System in ready state.',
            steps=[
                'Step 1: Trigger Update PortOut API with mismatched LineId and AccountId',
                'Step 2: Validate MBO throws error ERR161 - accountNumber and lineId mismatch',
            ],
            validation='ERR161 - accountNumber and lineId mismatch.',
            category='Negative',
        ),
        ChalkScenario(
            scenario_id='TS_MWTGPROV-3941_22',
            title='Update PortOut - requestNumber Mismatch',
            prereq='Pre-req: System in ready state.',
            steps=[
                'Step 1: Trigger Update PortOut API with invalid requestNumber',
                'Step 2: Validate MBO throws error for requestNumber mismatch',
            ],
            validation='MBO throws error for requestNumber mismatch.',
            category='Negative',
        ),
        ChalkScenario(
            scenario_id='TS_MWTGPROV-3941_23',
            title='Port-Out - Validate Transaction and Notification status if Open Order exists',
            prereq='Pre-req: Active TMO subscriber line with existing open order.',
            steps=[
                'Step 1: Trigger UP but not PO',
                'Step 2: Trigger next UP - NSL checks open order exists',
                'Step 3: Validate NSL updates transaction status accordingly',
            ],
            validation='NSL detects open order and updates transaction status.',
            category='Edge Case',
        ),
        ChalkScenario(
            scenario_id='TS_MWTGPROV-3941_24',
            title='Port-Out - Validate status if notification not sent to ITMBO or any error',
            prereq='Pre-req: Active TMO subscriber line.',
            steps=[
                'Step 1: Trigger Unsolicited PortOut API',
                'Step 2: Simulate notification failure to ITMBO',
                'Step 3: Validate Transaction and Notification status reflect failure',
            ],
            validation='Notification failure handled. Status updated to Failed.',
            category='Negative',
        ),
        ChalkScenario(
            scenario_id='TS_MWTGPROV-3941_25',
            title='Port-Out - Validate status if PO nsl-updateportout-tmo receives any error',
            prereq='Pre-req: Active TMO subscriber line. UP completed.',
            steps=[
                'Step 1: Trigger UP with correct MDN',
                'Step 2: PO triggered but nsl-updateportout-tmo receives error',
                'Step 3: Validate Transaction and Notification status',
            ],
            validation='PO error handled. Transaction status updated.',
            category='Negative',
        ),
    ],
    open_items=[
        'NBOP UI validation for port-out: Change MDN/Device and SIM options should not display for ported-out line',
        'Apollo-NE portal validation: NA at this moment for this iteration',
    ],
)

# ================================================================
# BUILD TEST SUITE
# ================================================================
print('=' * 80)
print('DRY RUN: MWTGPROV-3941 Port-Out')
print('=' * 80)

options = {
    'channel': ['ITMBO'],
    'devices': ['Mobile'],
    'networks': ['4G', '5G'],
    'sim_types': ['eSIM', 'pSIM'],
    'os_platforms': ['iOS', 'Android'],
    'include_positive': True,
    'include_negative': True,
    'include_e2e': True,
    'include_edge': True,
    'include_attachments': False,
    'strategy': 'Smart Suite (Recommended)',
    'custom_instructions': '',
}

suite = build_test_suite(jira, chalk, [], options, log=print)
out_path = generate_excel(suite, log=print)

# ================================================================
# COMPARE WITH SAMPLE
# ================================================================
print('\n' + '=' * 80)
print('COMPARISON WITH SAMPLE')
print('=' * 80)

import openpyxl
sample_path = os.path.join(os.path.dirname(__file__), '..', 'Samples', 'testcases_1775668767931_PgZHORNLN7(1).xlsx')
swb = openpyxl.load_workbook(sample_path, data_only=True)
sws = swb['TestCases']

# Extract sample TCs for 3941
sample_tcs = []
for r in range(2, sws.max_row + 1):
    summary = str(sws.cell(r, 1).value or '')
    labels = str(sws.cell(r, 6).value or '')
    if '3941' in summary or '3941' in labels:
        if summary not in [s['summary'] for s in sample_tcs]:
            sample_tcs.append({
                'summary': summary,
                'desc': str(sws.cell(r, 2).value or '')[:150],
            })

print('\nSample TCs: %d' % len(sample_tcs))
print('Generated TCs: %d' % len(suite.test_cases))
total_steps = sum(len(tc.steps) for tc in suite.test_cases)
print('Generated Steps: %d' % total_steps)

# Category breakdown
cats = {}
for tc in suite.test_cases:
    cats.setdefault(tc.category, 0)
    cats[tc.category] += 1
print('\nCategory Breakdown:')
for cat, count in sorted(cats.items()):
    print('  %s: %d' % (cat, count))

# Coverage analysis: check which sample scenarios are covered
print('\n--- SAMPLE SCENARIO COVERAGE ---')
sample_keywords = {}
for stc in sample_tcs:
    name = stc['summary']
    # Extract key concepts
    kws = set()
    name_low = name.lower()
    if 'hotline' in name_low: kws.add('hotline')
    if 'suspend' in name_low: kws.add('suspend')
    if 'deactiv' in name_low: kws.add('deactivated')
    if 'mdn_less' in name_low or 'less_than_10' in name_low: kws.add('mdn_length')
    if 'mdn_not_exists' in name_low: kws.add('mdn_not_found')
    if 'invalid_pin' in name_low or '6a' in name_low: kws.add('invalid_pin')
    if 'invalid account' in name_low or '8a' in name_low: kws.add('invalid_account')
    if 'mismatch' in name_low: kws.add('mismatch')
    if 'requestnumber' in name_low: kws.add('request_number')
    if 'cancel' in name_low: kws.add('cancel')
    if 'notification' in name_low and 'failed' in name_low: kws.add('notification_failed')
    if '50_secs' in name_low or '50 secs' in name_low: kws.add('timeout_50s')
    if 'rescheduled' in name_low: kws.add('rescheduled')
    if 'secondary' in name_low or 'sec environment' in name_low: kws.add('secondary_env')
    if 'open order' in name_low: kws.add('open_order')
    if 'without_transaction' in name_low: kws.add('missing_fields')
    if 'esim' in name_low: kws.add('esim')
    if 'psim' in name_low: kws.add('psim')
    if 'up_po_success' in name_low or 'up_po_succ' in name_low: kws.add('happy_path')
    if 'up_po_failure' in name_low: kws.add('po_failure')
    if 'update portout' in name_low and 'invalid' in name_low: kws.add('update_po_validation')
    sample_keywords[name[:80]] = kws

# Check generated TCs
gen_text = ' '.join([
    tc.summary.lower() + ' ' + tc.description.lower() + ' ' +
    ' '.join(s.summary.lower() for s in tc.steps)
    for tc in suite.test_cases
])

covered = set()
for name, kws in sample_keywords.items():
    for kw in kws:
        if kw.replace('_', ' ') in gen_text or kw in gen_text:
            covered.add(kw)

all_kws = set()
for kws in sample_keywords.values():
    all_kws.update(kws)

print('\nScenario concepts in sample: %d' % len(all_kws))
print('Covered by generated suite: %d' % len(covered))
missing = all_kws - covered
if missing:
    print('MISSING concepts: %s' % ', '.join(sorted(missing)))
else:
    print('ALL sample concepts covered!')

print('\n--- GENERATED TC LIST ---')
for tc in suite.test_cases:
    print('  %s | %s | %d steps' % (tc.sno.zfill(2), tc.summary[:90], len(tc.steps)))

print('\nOutput: %s' % out_path)
print('Done!')
