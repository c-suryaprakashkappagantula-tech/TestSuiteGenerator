"""
Dry-run: Build test suite for MWTGPROV-3941 from DB cache, generate Excel, compare with sample.
"""
import sys, os, json
sys.path.insert(0, os.path.dirname(__file__))

from modules.config import OUTPUTS
from modules.test_engine import build_test_suite
from modules.jira_fetcher import JiraIssue
from modules.chalk_parser import ChalkData, ChalkScenario
from modules.excel_generator import generate_excel
from modules.database import load_chalk_as_object, _conn

# ================================================================
# LOAD CHALK FROM DB
# ================================================================
print('Loading MWTGPROV-3941 from DB cache...')

# Try all PIs
c = _conn()
row = c.execute("SELECT pi_label FROM chalk_cache WHERE feature_id='MWTGPROV-3941' AND scenarios_json != '[]' LIMIT 1").fetchone()
c.close()

if row:
    chalk = load_chalk_as_object('MWTGPROV-3941', row['pi_label'])
    print('Found in DB under %s: %d scenarios' % (row['pi_label'], len(chalk.scenarios)))
    for i, s in enumerate(chalk.scenarios, 1):
        print('  %2d. [%-10s] %s' % (i, s.category, s.title[:80]))
else:
    print('NOT FOUND in DB! Run preload_db.py first.')
    sys.exit(1)

# ================================================================
# MOCK JIRA (same as before)
# ================================================================
jira = JiraIssue(
    key='MWTGPROV-3941',
    summary='Port-Out - Unsolicited Port Out / Update Port Out / Cancel Port Out',
    description='Port-Out feature for TMO integration. NSL processes UP/PO/Cancel requests.',
    status='In Progress', priority='High', issue_type='Epic',
    assignee='QA Team', reporter='Dev Lead',
    labels=['MWTGPROV-3941', 'PI-51', 'PortOut', 'TMO'],
    pi='PI-51', channel='ITMBO',
    acceptance_criteria="""
* NSL shall process Unsolicited Port-Out (UP) from TMO via NE and forward to IT-MBO
* NSL shall process Update Port-Out (PO) from MBO and trigger deactivation
* NSL shall handle Cancel Port-Out and update MNP status to Cancelled
* NSL shall reject port-out for Hotlined, Suspended, and Deactivated MDNs
* NSL shall validate MDN exists and is in Active status
* NSL shall handle IT-MBO rejection responses (Invalid PIN 6A, Invalid Account 8A)
* NSL shall update Transaction History and MNP_DETAILS tables
* NSL shall trigger port-out deactivation within 1 minute of successful PO
* NSL shall verify all backend calls in Century Report
""",
)

# ================================================================
# BUILD
# ================================================================
print('\n' + '=' * 60)
print('BUILDING TEST SUITE')
print('=' * 60)

options = {
    'channel': ['ITMBO'], 'devices': ['Mobile'], 'networks': ['4G', '5G'],
    'sim_types': ['eSIM', 'pSIM'], 'os_platforms': ['iOS', 'Android'],
    'include_positive': True, 'include_negative': True, 'include_e2e': True,
    'include_edge': True, 'include_attachments': False,
    'strategy': 'Smart Suite (Recommended)', 'custom_instructions': '',
}

suite = build_test_suite(jira, chalk, [], options, log=print)
out_path = generate_excel(suite, log=print)

# ================================================================
# COMPARE WITH SAMPLE
# ================================================================
print('\n' + '=' * 60)
print('COMPARISON WITH SAMPLE')
print('=' * 60)

import openpyxl

# Our output
print('\n--- OUR OUTPUT ---')
owb = openpyxl.load_workbook(str(out_path), data_only=True)
our_tcs = {}
for sn in owb.sheetnames:
    if sn in ('Summary', 'Traceability', 'Combinations'): continue
    ws = owb[sn]
    count = 0
    for r in range(3, ws.max_row + 1):
        if ws.cell(r, 1).value and str(ws.cell(r, 1).value).strip():
            count += 1
    our_tcs[sn] = count
    print('  %-35s %3d TCs' % (sn, count))
our_total = sum(our_tcs.values())
print('  %-35s %3d TCs' % ('TOTAL', our_total))

# Sample
print('\n--- SAMPLE (3941 only) ---')
swb = openpyxl.load_workbook(os.path.join(os.path.dirname(__file__), '..', 'Samples', 'testcases_1775668767931_PgZHORNLN7(1).xlsx'), data_only=True)
ws = swb['TestCases']
sample_tcs = set()
sample_summaries = []
for r in range(2, ws.max_row + 1):
    summary = str(ws.cell(r, 1).value or '')
    labels = str(ws.cell(r, 6).value or '')
    if '3941' in summary or '3941' in labels:
        if summary not in sample_tcs:
            sample_tcs.add(summary)
            sample_summaries.append(summary[:100])
print('  Total unique TCs: %d' % len(sample_tcs))

# Scenario coverage
print('\n--- SCENARIO COVERAGE ---')
our_text = ''
for sn in owb.sheetnames:
    if sn in ('Summary', 'Traceability', 'Combinations'): continue
    ws = owb[sn]
    for r in range(3, ws.max_row + 1):
        for c in range(1, 8):
            v = str(ws.cell(r, c).value or '')
            our_text += ' ' + v.lower()

checks = {
    'Port-Out UP happy path': ['port out', 'unsolicited'],
    'Port-Out PO/Update': ['update port', 'po '],
    'Cancel Port-Out': ['cancel'],
    'Hotlined MDN rejected': ['hotline'],
    'Suspended MDN rejected': ['suspend'],
    'Deactivated MDN rejected': ['deactiv'],
    'MDN not found (ERR07)': ['mdn not found', 'err07'],
    'Schema validation errors': ['schema validation'],
    'Invalid LineId': ['invalid lineid', 'non-existent lineid'],
    'Invalid AccountId': ['invalid accountid', 'non-existent accountid'],
    'LineId/AccountId mismatch': ['lineid and accountid', 'mismatch'],
    'Duplicate UP requests': ['duplicate'],
    'Port-out already completed': ['already completed', 'already ported'],
    'Rollback': ['rollback'],
    'Transaction History': ['transaction history'],
    'Century Report': ['century report', 'service grouping'],
    'NBOP Portal': ['nbop', 'portal'],
    'E2E flow': ['end-to-end', 'e2e'],
    'Invalid MDN format': ['invalid mdn', 'less than 10'],
    'Upstream rejection': ['upstream rejection', 'itmbo rejection'],
}

covered = 0
for name, keywords in checks.items():
    found = any(kw in our_text for kw in keywords)
    if found: covered += 1
    print('  %-35s %s' % (name, 'YES' if found else 'MISSING'))

print('\n  Coverage: %d/%d (%.0f%%)' % (covered, len(checks), covered/len(checks)*100))

# TC name quality check
print('\n--- TC NAME QUALITY ---')
bad_names = 0
for sn in owb.sheetnames:
    if sn in ('Summary', 'Traceability', 'Combinations'): continue
    ws = owb[sn]
    for r in range(3, ws.max_row + 1):
        summ = str(ws.cell(r, 2).value or '')
        if summ and len(summ) > 100:
            bad_names += 1
            if bad_names <= 3:
                print('  LONG: %s...' % summ[:80])
print('  TCs with summary > 100 chars: %d' % bad_names)

print('\n--- SAMPLE TC NAMES (first 10) ---')
for s in sample_summaries[:10]:
    print('  %s' % s)

print('\nOutput: %s' % out_path)
