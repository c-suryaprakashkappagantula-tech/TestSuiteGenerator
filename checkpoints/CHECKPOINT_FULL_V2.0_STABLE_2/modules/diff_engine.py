"""
diff_engine.py — V2.1: Compare two test suite Excel files and produce a diff report.
Identifies new, changed, and removed test cases between versions.
"""
import re
from pathlib import Path
from typing import Dict, List
import openpyxl


def compare_suites(old_path: Path, new_path: Path, feature_id: str, log=print) -> Dict:
    """Compare old and new test suite Excel files.
    Returns dict with: new, changed, removed counts + details list."""
    log('[DIFF] Loading old suite: %s' % Path(old_path).name)
    old_tcs = _extract_tcs_from_excel(old_path, feature_id)
    log('[DIFF] Loading new suite: %s' % Path(new_path).name)
    new_tcs = _extract_tcs_from_excel(new_path, feature_id)

    log('[DIFF] Old: %d TCs | New: %d TCs' % (len(old_tcs), len(new_tcs)))

    # Build keyword fingerprints for fuzzy matching
    old_fps = {name: _fingerprint(tc) for name, tc in old_tcs.items()}
    new_fps = {name: _fingerprint(tc) for name, tc in new_tcs.items()}

    # Match by fingerprint similarity
    matched_old = set()
    matched_new = set()
    changed = []
    details = []

    for new_name, new_fp in new_fps.items():
        best_match = None
        best_score = 0
        for old_name, old_fp in old_fps.items():
            if old_name in matched_old:
                continue
            score = _similarity(new_fp, old_fp)
            if score > best_score:
                best_score = score
                best_match = old_name

        if best_match and best_score >= 0.5:
            matched_old.add(best_match)
            matched_new.add(new_name)
            if best_score < 0.9:
                changed.append(new_name)
                details.append('CHANGED: %s (was: %s, similarity: %.0f%%)' % (
                    new_name[:70], best_match[:70], best_score * 100))

    # New TCs = in new but not matched
    new_only = [n for n in new_fps if n not in matched_new]
    for n in new_only:
        details.append('NEW: %s' % n[:90])

    # Removed TCs = in old but not matched
    removed_only = [o for o in old_fps if o not in matched_old]
    for o in removed_only:
        details.append('REMOVED: %s' % o[:90])

    report = {
        'new': len(new_only),
        'changed': len(changed),
        'removed': len(removed_only),
        'matched': len(matched_new) - len(changed),
        'old_total': len(old_tcs),
        'new_total': len(new_tcs),
        'details': details,
    }

    log('[DIFF] Result: %d new, %d changed, %d removed, %d unchanged' % (
        report['new'], report['changed'], report['removed'], report['matched']))
    return report


def _extract_tcs_from_excel(path: Path, feature_id: str) -> Dict[str, Dict]:
    """Extract TC summaries and descriptions from an Excel file."""
    tcs = {}
    try:
        wb = openpyxl.load_workbook(str(path), data_only=True)
        for sn in wb.sheetnames:
            ws = wb[sn]
            for r in range(1, ws.max_row + 1):
                # Try column A for summary
                val = str(ws.cell(r, 1).value or '')
                if not val or len(val) < 10:
                    continue
                # Check if this row is a TC (contains feature ID or TC pattern)
                if feature_id in val.upper() or re.match(r'TC\d+', val):
                    desc = str(ws.cell(r, 2).value or '') + ' ' + str(ws.cell(r, 3).value or '')
                    steps_text = ''
                    # Collect step text from subsequent rows or columns
                    for sc in range(5, min(ws.max_column + 1, 10)):
                        sv = str(ws.cell(r, sc).value or '')
                        if sv:
                            steps_text += ' ' + sv
                    tcs[val[:120]] = {
                        'summary': val,
                        'description': desc[:300],
                        'steps': steps_text[:500],
                    }
    except Exception as e:
        pass  # Return empty if file can't be parsed
    return tcs


def _fingerprint(tc: Dict) -> set:
    """Create a keyword fingerprint for a TC."""
    text = (tc.get('summary', '') + ' ' + tc.get('description', '') + ' ' + tc.get('steps', '')).lower()
    return set(re.findall(r'\b\w{4,}\b', text))


def _similarity(fp1: set, fp2: set) -> float:
    """Jaccard similarity between two fingerprints."""
    if not fp1 or not fp2:
        return 0.0
    return len(fp1 & fp2) / len(fp1 | fp2)
