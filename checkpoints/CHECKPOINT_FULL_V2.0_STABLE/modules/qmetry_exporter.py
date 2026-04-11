"""
qmetry_exporter.py — V2.1: Generate QMetry-compatible Excel for direct import.
Matches the exact column structure QMetry expects:
Summary, Description, Precondition, Status, Priority, Labels,
Step Summary, Expected Result, Version, Folders, TestCase Type, Type,
References, Scenario, Source TestCase ID, Steps, Section
"""
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from .config import OUTPUTS, ts
from .test_engine import TestSuite

QMETRY_HEADERS = [
    'Summary', 'Description', 'Precondition', 'Status', 'Priority',
    'Labels', 'Step Summary', 'Expected Result', 'Version', 'Folders',
    'TestCase Type', 'Type', 'References', 'Scenario', 'Source TestCase ID',
    'Steps', 'Section',
]

_hf = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
_hfill = PatternFill(start_color='0B1D39', end_color='0B1D39', fill_type='solid')
_nf = Font(name='Calibri', size=10)
_wrap = Alignment(wrap_text=True, vertical='top')
_bdr = Border(left=Side(style='thin'), right=Side(style='thin'),
              top=Side(style='thin'), bottom=Side(style='thin'))


def generate_qmetry_excel(suite: TestSuite, log=print, folder_path='') -> Path:
    """Generate QMetry-compatible Excel. Returns output path."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'TestCases'

    # Column widths
    widths = [60, 50, 40, 10, 10, 25, 60, 50, 6, 60, 12, 10, 20, 20, 20, 6, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else 'A' + chr(64 + i - 26)].width = w

    # Headers
    for ci, h in enumerate(QMETRY_HEADERS, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = _hf; c.fill = _hfill; c.alignment = _wrap; c.border = _bdr
    ws.freeze_panes = 'A2'

    row = 2
    for tc in suite.test_cases:
        # Build labels: feature_id + category
        labels = suite.feature_id
        if tc.category and tc.category != 'Happy Path':
            labels += '_%s' % tc.category.replace(' ', '_')

        # Confidence tag in description if available
        desc = tc.description
        if hasattr(tc, 'confidence') and tc.confidence:
            desc = '[%s] %s' % (tc.confidence, desc)

        if not tc.steps:
            # Single row, no steps
            _write_qmetry_row(ws, row, tc, labels, desc, '', '', suite, folder_path)
            row += 1
        else:
            # First step row gets the TC-level fields
            for si, step in enumerate(tc.steps):
                if si == 0:
                    _write_qmetry_row(ws, row, tc, labels, desc,
                                      step.summary, step.expected, suite, folder_path)
                else:
                    # Subsequent step rows: only Step Summary + Expected Result
                    ws.cell(row=row, column=7, value=step.summary).font = _nf
                    ws.cell(row=row, column=7).alignment = _wrap
                    ws.cell(row=row, column=7).border = _bdr
                    ws.cell(row=row, column=8, value=step.expected).font = _nf
                    ws.cell(row=row, column=8).alignment = _wrap
                    ws.cell(row=row, column=8).border = _bdr
                    # Empty cells still need borders
                    for ci in [1,2,3,4,5,6,9,10,11,12,13,14,15,16,17]:
                        ws.cell(row=row, column=ci).border = _bdr
                row += 1

    out = OUTPUTS / ('QMETRY_%s_%s.xlsx' % (suite.feature_id, ts()))
    wb.save(str(out))
    log('[QMETRY] Saved: %s (%d TCs, %d rows)' % (out.name, len(suite.test_cases), row - 2))
    return out


def _write_qmetry_row(ws, row, tc, labels, desc, step_sum, step_exp, suite, folder_path):
    """Write a single QMetry row with all TC-level fields."""
    ws.cell(row=row, column=1, value=tc.summary).font = _nf           # Summary
    ws.cell(row=row, column=1).alignment = _wrap; ws.cell(row=row, column=1).border = _bdr
    ws.cell(row=row, column=2, value=desc).font = _nf                 # Description
    ws.cell(row=row, column=2).alignment = _wrap; ws.cell(row=row, column=2).border = _bdr
    ws.cell(row=row, column=3, value=tc.preconditions).font = _nf     # Precondition
    ws.cell(row=row, column=3).alignment = _wrap; ws.cell(row=row, column=3).border = _bdr
    ws.cell(row=row, column=4).border = _bdr                          # Status (empty)
    ws.cell(row=row, column=5).border = _bdr                          # Priority (empty)
    ws.cell(row=row, column=6, value=labels).font = _nf               # Labels
    ws.cell(row=row, column=6).border = _bdr
    ws.cell(row=row, column=7, value=step_sum).font = _nf             # Step Summary
    ws.cell(row=row, column=7).alignment = _wrap; ws.cell(row=row, column=7).border = _bdr
    ws.cell(row=row, column=8, value=step_exp).font = _nf             # Expected Result
    ws.cell(row=row, column=8).alignment = _wrap; ws.cell(row=row, column=8).border = _bdr
    ws.cell(row=row, column=9, value='1').font = _nf                  # Version
    ws.cell(row=row, column=9).border = _bdr
    ws.cell(row=row, column=10, value=folder_path).font = _nf         # Folders
    ws.cell(row=row, column=10).border = _bdr
    ws.cell(row=row, column=11, value='Manual').font = _nf            # TestCase Type
    ws.cell(row=row, column=11).border = _bdr
    ws.cell(row=row, column=12).border = _bdr                         # Type (empty)
    ws.cell(row=row, column=13, value=tc.story_linkage).font = _nf    # References
    ws.cell(row=row, column=13).border = _bdr
    ws.cell(row=row, column=14).border = _bdr                         # Scenario
    ws.cell(row=row, column=15).border = _bdr                         # Source TestCase ID
    ws.cell(row=row, column=16).border = _bdr                         # Steps
    ws.cell(row=row, column=17).border = _bdr                         # Section
