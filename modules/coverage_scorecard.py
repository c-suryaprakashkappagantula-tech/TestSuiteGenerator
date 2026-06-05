# -*- coding: utf-8 -*-
"""
coverage_scorecard.py — Coverage Scorecard for V8.0 Test Suites.

Computes and formats a Coverage Scorecard — a single headline artifact
that answers "is this suite good enough?" across multiple dimensions.

Lenses:
  1. AC coverage %    — how many Jira ACs have covering TCs
  2. Business rules   — how many NMNO error codes have negative TCs
  3. Line-state matrix — how many of 7 states have TCs
  4. Downstream systems — must_call covered? must_not_call asserted?
  5. Chalk scenarios  — how many Chalk scenarios were mapped to TCs
  6. Grounding %      — mean grounding score across all TCs
  7. Category balance — Pos:Neg:Edge ratio
  8. Redundancy %     — deduplication signal

Usage:
    from modules.coverage_scorecard import compute_scorecard, format_scorecard_text

    scorecard = compute_scorecard(suite, jira, contract=None, nmno_result=None)
    print(format_scorecard_text(scorecard))
"""

from dataclasses import dataclass, field
from typing import List, Dict, Optional, Any
import re


# Line states the matrix should cover
_STATE_MATRIX = [
    'active', 'suspended', 'hotlined',
    'pending port-out', 'pending port-in', 'cancelled', 'pre-active',
]


@dataclass
class CoverageLens:
    """A single coverage measurement."""
    name: str
    covered: int
    total: int
    gaps: List[str] = field(default_factory=list)
    risk: str = ''   # 'HIGH', 'MEDIUM', 'LOW', ''

    @property
    def pct(self) -> float:
        return round(100.0 * self.covered / self.total, 1) if self.total > 0 else 100.0

    @property
    def badge(self) -> str:
        p = self.pct
        return '🟢' if p >= 80 else ('🟡' if p >= 50 else '🔴')

    @property
    def display(self) -> str:
        if self.total == 0:
            return 'N/A'
        return '%d/%d (%.0f%%)' % (self.covered, self.total, self.pct)


@dataclass
class CoverageScorecard:
    """Complete coverage scorecard for a test suite."""
    feature_id: str
    feature_title: str
    lenses: List[CoverageLens] = field(default_factory=list)
    grounding_pct: float = -1.0
    redundancy_pct: float = 0.0
    overall_risk: str = ''   # 'HIGH', 'MEDIUM', 'LOW'
    gaps_summary: List[str] = field(default_factory=list)

    @property
    def headline_badge(self) -> str:
        if self.overall_risk == 'HIGH':
            return '🔴'
        elif self.overall_risk == 'MEDIUM':
            return '🟡'
        return '🟢'


def compute_scorecard(
    suite,
    jira=None,
    contract=None,
    nmno_result=None,
    log=None,
) -> CoverageScorecard:
    """Compute the coverage scorecard for a test suite.

    Args:
        suite: TestSuite object with test_cases
        jira: JiraIssue object (for AC coverage)
        contract: OperationContract (for downstream system coverage)
        nmno_result: NMNOLookupResult (for business rule coverage)
        log: Logger function

    Returns:
        CoverageScorecard with all lenses computed.
    """
    log = log or print
    feature_id = getattr(suite, 'feature_id', '') or ''
    feature_title = getattr(suite, 'feature_title', '') or ''
    tcs = suite.test_cases or []

    # Build a searchable text from all TC summaries + step text
    _all_tc_text = ' '.join(
        (tc.summary or '') + ' ' + ' '.join(
            (s.summary or '') + ' ' + (s.expected or '') for s in (tc.steps or [])
        )
        for tc in tcs
    ).lower()

    lenses: List[CoverageLens] = []
    gaps_summary: List[str] = []

    # ── 1. AC Coverage ──
    ac_items = []
    if jira and hasattr(jira, 'acceptance_criteria') and jira.acceptance_criteria:
        _ac = jira.acceptance_criteria
        if isinstance(_ac, str) and _ac.strip():
            ac_items = [line.strip(' *-#') for line in _ac.split('\n')
                        if len(line.strip()) > 15]
        elif isinstance(_ac, list):
            ac_items = [str(a) for a in _ac if len(str(a)) > 15]

    if ac_items:
        covered_acs = 0
        uncovered_acs = []
        for ac in ac_items:
            # Extract key words from AC and check if any TC covers them
            ac_words = set(re.findall(r'\b\w{4,}\b', ac.lower()))
            ac_words -= {'shall', 'must', 'should', 'verify', 'ensure', 'when', 'then', 'given'}
            if ac_words:
                overlap = sum(1 for w in ac_words if w in _all_tc_text) / len(ac_words)
                if overlap >= 0.4:
                    covered_acs += 1
                else:
                    uncovered_acs.append(ac[:80])
            else:
                covered_acs += 1  # trivially covered (too short to judge)

        risk = 'HIGH' if len(uncovered_acs) >= 2 else ('MEDIUM' if uncovered_acs else 'LOW')
        lens = CoverageLens(
            name='AC coverage',
            covered=covered_acs,
            total=len(ac_items),
            gaps=uncovered_acs[:3],
            risk=risk,
        )
        lenses.append(lens)
        if uncovered_acs:
            gaps_summary.append('%d uncovered AC(s) — %s' % (len(uncovered_acs), risk))

    # ── 2. Business Rule (NMNO error code) coverage ──
    if nmno_result and nmno_result.business_rules:
        total_rules = len(nmno_result.business_rules)
        covered_rules = 0
        missing_codes = []
        for rule in nmno_result.business_rules:
            ec = (rule.error_code or '').upper()
            if ec and ec in _all_tc_text.upper():
                covered_rules += 1
            elif ec:
                missing_codes.append(ec)
            else:
                covered_rules += 1  # no code to check
        risk = 'HIGH' if len(missing_codes) > 3 else ('MEDIUM' if missing_codes else 'LOW')
        lens = CoverageLens(
            name='Business rules',
            covered=covered_rules,
            total=total_rules,
            gaps=missing_codes[:5],
            risk=risk,
        )
        lenses.append(lens)
        if missing_codes:
            gaps_summary.append('Missing error codes: %s' % ', '.join(missing_codes[:5]))

    # ── 3. Line-state matrix coverage ──
    covered_states = 0
    missing_states = []
    for state in _STATE_MATRIX:
        if state in _all_tc_text:
            covered_states += 1
        else:
            missing_states.append(state.title())
    risk = 'HIGH' if len(missing_states) >= 4 else ('MEDIUM' if missing_states else 'LOW')
    lens = CoverageLens(
        name='Line-state matrix',
        covered=covered_states,
        total=len(_STATE_MATRIX),
        gaps=missing_states[:4],
        risk=risk,
    )
    lenses.append(lens)
    if missing_states:
        gaps_summary.append('Line states not tested: %s' % ', '.join(missing_states[:3]))

    # ── 4. Downstream system coverage ──
    if contract:
        must_call = contract.must_call or []
        must_not_call = contract.must_not_call or []
        total_sys = len(must_call) + len(must_not_call)
        covered_sys = 0
        missing_sys = []

        for sys_key in must_call:
            sys_name = sys_key.replace('_', ' ')
            if sys_name in _all_tc_text or sys_key in _all_tc_text:
                covered_sys += 1
            else:
                missing_sys.append('%s (must-call)' % sys_name.upper())

        for sys_key in must_not_call:
            sys_name = sys_key.replace('_', ' ')
            if ('not call' in _all_tc_text and sys_name in _all_tc_text) or \
               ('not called' in _all_tc_text and sys_name in _all_tc_text) or \
               ('must not' in _all_tc_text and sys_name in _all_tc_text):
                covered_sys += 1
            else:
                missing_sys.append('%s (must-not-call)' % sys_name.upper())

        if total_sys > 0:
            risk = 'MEDIUM' if missing_sys else 'LOW'
            lens = CoverageLens(
                name='Downstream systems',
                covered=covered_sys,
                total=total_sys,
                gaps=missing_sys[:4],
                risk=risk,
            )
            lenses.append(lens)

    # ── 5. Chalk scenario completeness ──
    if hasattr(suite, 'data_inventory') and suite.data_inventory:
        _chalk_source = next(
            (s for s in suite.data_inventory.sources if s.source_type == 'chalk' and
             'scenario' in s.source_name.lower()),
            None
        )
        if _chalk_source and _chalk_source.items_extracted > 0:
            # How many Chalk scenarios map to TCs?
            chalk_total = _chalk_source.items_extracted
            # Estimate: count TCs sourced from Chalk
            chalk_covered = sum(
                1 for tc in tcs
                if tc.traceability and
                   getattr(tc.traceability, 'source_type', '') in ('Chalk Scenario',)
            )
            if chalk_covered == 0:
                # Fallback: TCs count that aren't negative/partial-failure
                chalk_covered = sum(
                    1 for tc in tcs
                    if tc.category not in ('Negative',) and 'partial' not in (tc.summary or '').lower()
                )
                chalk_covered = min(chalk_covered, chalk_total)
            lens = CoverageLens(
                name='Chalk scenarios',
                covered=chalk_covered,
                total=chalk_total,
                risk='MEDIUM' if chalk_covered < chalk_total else 'LOW',
            )
            lenses.append(lens)

    # ── 6. Category balance ──
    cat_counts = {}
    for tc in tcs:
        cat = tc.category or 'Happy Path'
        cat_counts[cat] = cat_counts.get(cat, 0) + 1
    pos = cat_counts.get('Happy Path', 0) + cat_counts.get('Positive', 0)
    neg = cat_counts.get('Negative', 0)
    edge = cat_counts.get('Edge Case', 0)
    total_tcs = len(tcs)
    if total_tcs > 0:
        neg_ratio = neg / total_tcs
        _cat_risk = 'HIGH' if neg_ratio < 0.15 else ('MEDIUM' if neg_ratio < 0.25 else 'LOW')
        lens = CoverageLens(
            name='Category balance',
            covered=neg,
            total=total_tcs,
            gaps=['Only %d negative TCs (%.0f%%)' % (neg, neg_ratio * 100)] if _cat_risk != 'LOW' else [],
            risk=_cat_risk,
        )
        lenses.append(lens)

    # ── 7. Grounding % ──
    try:
        from .grounding_scorer import suite_grounding_pct, grounding_badge
        grounding_pct = suite_grounding_pct(tcs)
    except Exception:
        grounding_pct = -1.0

    # ── 8. Redundancy ──
    summaries = [tc.summary or '' for tc in tcs]
    if summaries:
        unique_summaries = len(set(s.lower()[:60] for s in summaries))
        redundancy_pct = round(100.0 * (1.0 - unique_summaries / len(summaries)), 1)
    else:
        redundancy_pct = 0.0

    # ── Overall risk ──
    high_count = sum(1 for l in lenses if l.risk == 'HIGH')
    medium_count = sum(1 for l in lenses if l.risk == 'MEDIUM')
    if high_count >= 1:
        overall_risk = 'HIGH'
    elif medium_count >= 2:
        overall_risk = 'MEDIUM'
    else:
        overall_risk = 'LOW'

    return CoverageScorecard(
        feature_id=feature_id,
        feature_title=feature_title,
        lenses=lenses,
        grounding_pct=grounding_pct,
        redundancy_pct=redundancy_pct,
        overall_risk=overall_risk,
        gaps_summary=gaps_summary,
    )


def format_scorecard_text(sc: CoverageScorecard) -> str:
    """Format scorecard as a readable text block."""
    lines = [
        'COVERAGE SCORECARD — %s' % sc.feature_id,
        ('  %s' % sc.feature_title[:70]) if sc.feature_title else '',
        '─' * 55,
    ]
    for lens in sc.lenses:
        gap_str = ''
        if lens.gaps:
            gap_str = ' — gaps: %s' % ', '.join(str(g)[:40] for g in lens.gaps[:2])
        lines.append('  %s %-22s %s%s' % (
            lens.badge, lens.name[:22], lens.display, gap_str[:50]))

    lines.append('─' * 55)
    if sc.grounding_pct >= 0:
        from .grounding_scorer import grounding_badge
        lines.append('  %s %-22s %.1f%%' % (grounding_badge(sc.grounding_pct), 'Grounding', sc.grounding_pct))
    if sc.redundancy_pct > 5:
        lines.append('  ⚠️  %-22s %.1f%%' % ('Redundancy', sc.redundancy_pct))
    lines.append('─' * 55)
    lines.append('  Overall risk: %s %s' % (sc.headline_badge, sc.overall_risk))
    if sc.gaps_summary:
        lines.append('  Key gaps:')
        for g in sc.gaps_summary[:3]:
            lines.append('    • %s' % g)
    return '\n'.join(l for l in lines if l is not None)


def build_scorecard_excel_sheet(wb, sc: CoverageScorecard):
    """Add a Coverage Scorecard sheet to an existing openpyxl Workbook."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    ws = wb.create_sheet('Coverage Scorecard')
    ws.column_dimensions['A'].width = 26
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 40

    _hf = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    _hfill = PatternFill(start_color='1A237E', end_color='1A237E', fill_type='solid')
    _nf = Font(name='Calibri', size=11)
    _bf = Font(name='Calibri', bold=True, size=11)
    _wrap = Alignment(wrap_text=True, vertical='top')
    _center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    _bdr = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    # Title
    ws.merge_cells('A1:D1')
    ws.cell(row=1, column=1, value='COVERAGE SCORECARD — %s' % sc.feature_id)
    ws.cell(row=1, column=1).font = Font(name='Calibri', bold=True, size=14, color='1A237E')
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # Subtitle
    ws.merge_cells('A2:D2')
    ws.cell(row=2, column=1, value=sc.feature_title[:100])
    ws.cell(row=2, column=1).font = Font(name='Calibri', italic=True, size=10)
    ws.row_dimensions[2].height = 16

    # Header
    r = 4
    for ci, h in enumerate(['Coverage Lens', 'Covered', 'Total', 'Gaps / Notes'], 1):
        c = ws.cell(row=r, column=ci, value=h)
        c.font = _hf; c.fill = _hfill; c.alignment = _center; c.border = _bdr
    r += 1

    _risk_fills = {
        'HIGH':   PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
        'MEDIUM': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
        'LOW':    PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
        '':       PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid'),
    }

    for lens in sc.lenses:
        fill = _risk_fills.get(lens.risk, _risk_fills[''])
        ws.cell(row=r, column=1, value='%s %s' % (lens.badge, lens.name)).font = _bf
        ws.cell(row=r, column=1).fill = fill; ws.cell(row=r, column=1).border = _bdr
        ws.cell(row=r, column=2, value=str(lens.covered)).font = _nf
        ws.cell(row=r, column=2).alignment = _center; ws.cell(row=r, column=2).border = _bdr
        ws.cell(row=r, column=3, value=str(lens.total)).font = _nf
        ws.cell(row=r, column=3).alignment = _center; ws.cell(row=r, column=3).border = _bdr
        gaps_text = '; '.join(str(g)[:60] for g in lens.gaps[:3]) if lens.gaps else '—'
        ws.cell(row=r, column=4, value=gaps_text).font = _nf
        ws.cell(row=r, column=4).alignment = _wrap; ws.cell(row=r, column=4).border = _bdr
        r += 1

    # Grounding + Redundancy row
    r += 1
    if sc.grounding_pct >= 0:
        try:
            from .grounding_scorer import grounding_badge
            _gb = grounding_badge(sc.grounding_pct)
        except Exception:
            _gb = '🟡'
        ws.cell(row=r, column=1, value='%s Grounding' % _gb).font = _bf; ws.cell(row=r, column=1).border = _bdr
        ws.cell(row=r, column=2, value='%.1f%%' % sc.grounding_pct).font = _nf
        ws.cell(row=r, column=2).alignment = _center; ws.cell(row=r, column=2).border = _bdr
        ws.cell(row=r, column=3, value='100%').font = _nf
        ws.cell(row=r, column=3).alignment = _center; ws.cell(row=r, column=3).border = _bdr
        ws.cell(row=r, column=4, value='Mean grounding score across all TCs').font = _nf
        ws.cell(row=r, column=4).alignment = _wrap; ws.cell(row=r, column=4).border = _bdr
        r += 1

    if sc.redundancy_pct > 0:
        ws.cell(row=r, column=1, value='⚠️ Redundancy').font = _bf; ws.cell(row=r, column=1).border = _bdr
        ws.cell(row=r, column=2, value='%.1f%%' % sc.redundancy_pct).font = _nf
        ws.cell(row=r, column=2).alignment = _center; ws.cell(row=r, column=2).border = _bdr
        ws.cell(row=r, column=3, value='0%').font = _nf
        ws.cell(row=r, column=3).alignment = _center; ws.cell(row=r, column=3).border = _bdr
        ws.cell(row=r, column=4, value='TC overlap signal — consider dedup').font = _nf
        ws.cell(row=r, column=4).alignment = _wrap; ws.cell(row=r, column=4).border = _bdr
        r += 1

    # Overall risk row
    r += 1
    _or_fill = _risk_fills.get(sc.overall_risk, _risk_fills[''])
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    ws.cell(row=r, column=1,
            value='%s OVERALL RISK: %s' % (sc.headline_badge, sc.overall_risk)).font = Font(
                name='Calibri', bold=True, size=13)
    ws.cell(row=r, column=1).fill = _or_fill
    ws.cell(row=r, column=1).alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[r].height = 24
