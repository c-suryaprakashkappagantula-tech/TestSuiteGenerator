"""
excel_generator.py — Generate production-ready Excel test suite.
Matches the exact format of TESTPLAN MWTGPROV-3976 sample.
3 sheets: Summary, Test Cases (merged+styled), Traceability.
"""
import shutil
from pathlib import Path
from typing import List
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from .config import (EXCEL_HEADERS, MERGE_COLS, NAVY, LIGHT_BLUE, WHITE,
                     CAT_COLORS, output_path, checkpoint_path)
from .test_engine import TestSuite, TestCase


# ── Shared styles ──
_hf = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
_hfill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type='solid')
_bf = Font(name='Calibri', bold=True, size=11, color=NAVY)
_nf = Font(name='Calibri', size=11)
_wrap = Alignment(wrap_text=True, vertical='top')
_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
_bdr = Border(left=Side(style='thin'), right=Side(style='thin'),
              top=Side(style='thin'), bottom=Side(style='thin'))

# ── Category row colors (very soft tints — easy on the eyes) ──
_CAT_ROW_FILLS = {
    'Happy Path':   PatternFill(start_color='F1F8E9', end_color='F1F8E9', fill_type='solid'),  # very light green
    'Positive':     PatternFill(start_color='F1F8E9', end_color='F1F8E9', fill_type='solid'),
    'Negative':     PatternFill(start_color='FBE9E7', end_color='FBE9E7', fill_type='solid'),  # very light coral
    'Edge Case':    PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid'),  # very light yellow
    'E2E':          PatternFill(start_color='E8EAF6', end_color='E8EAF6', fill_type='solid'),  # very light indigo
    'End-to-End':   PatternFill(start_color='E8EAF6', end_color='E8EAF6', fill_type='solid'),
    'Rollback':     PatternFill(start_color='F3E5F5', end_color='F3E5F5', fill_type='solid'),  # very light purple
}
_CAT_SNO_FONTS = {
    'Happy Path':   Font(name='Calibri', bold=True, size=11, color='388E3C'),   # medium green
    'Positive':     Font(name='Calibri', bold=True, size=11, color='388E3C'),
    'Negative':     Font(name='Calibri', bold=True, size=11, color='D32F2F'),   # medium red
    'Edge Case':    Font(name='Calibri', bold=True, size=11, color='F9A825'),   # amber
    'E2E':          Font(name='Calibri', bold=True, size=11, color='1976D2'),   # medium blue
    'End-to-End':   Font(name='Calibri', bold=True, size=11, color='1976D2'),
    'Rollback':     Font(name='Calibri', bold=True, size=11, color='7B1FA2'),   # medium purple
}
_lb = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type='solid')
_wf = PatternFill(start_color=WHITE, end_color=WHITE, fill_type='solid')


def generate_excel(suite: TestSuite, log=print) -> Path:
    """Generate the complete Excel workbook. Returns output path."""
    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ──
    log('[EXCEL] Building Summary sheet...')
    _build_summary_sheet(wb, suite)

    # ── Sheet 2: All Test Cases in ONE sheet ──
    log('[EXCEL] Building Test Cases sheet (%d TCs)...' % len(suite.test_cases))
    _build_testcases_sheet(wb, suite, sheet_name='Test Cases')

    # ── Sheet 3: Traceability (AC → TC mapping) ──
    if suite.ac_traceability:
        log('[EXCEL] Building Traceability sheet...')
        _build_traceability_sheet(wb, suite)

    # ── Sheet 4: Combinations (if matrix expansion was used) ──
    if hasattr(suite, 'combinations') and suite.combinations and len(suite.combinations) > 1:
        log('[EXCEL] Building Combinations sheet...')
        _build_combinations_sheet(wb, suite)

    # Remove default empty sheet if exists
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # Save
    out = output_path(suite.feature_id, pi=suite.pi, title=suite.feature_title)
    wb.save(str(out))
    log(f'[EXCEL] ✅ Saved: {out.name}')

    # Auto-checkpoint
    cp = checkpoint_path(suite.feature_id, pi=suite.pi, title=suite.feature_title)
    shutil.copy2(str(out), str(cp))
    log(f'[EXCEL] ✅ Checkpoint: {cp.name}')

    return out


def _build_summary_sheet(wb, suite: TestSuite):
    """Build the Summary sheet with feature info, AC, coverage, sources."""
    ws = wb.active
    ws.title = 'Summary'
    ws.freeze_panes = 'A2'  # Point 16: Freeze title row
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 58
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 10

    # Title
    r = 1
    ws.merge_cells(f'A{r}:F{r}')
    ws.cell(row=r, column=1, value=f'TEST PLAN SUMMARY - {suite.feature_id}')
    ws.cell(row=r, column=1).font = Font(name='Calibri', bold=True, size=16, color='1A237E')
    ws.cell(row=r, column=1).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=r, column=1).fill = PatternFill(start_color='E8EAF6', end_color='E8EAF6', fill_type='solid')
    ws.row_dimensions[r].height = 35
    r += 2

    # Helper for section headers
    _sec_font = Font(name='Calibri', bold=True, size=12, color='1A237E')
    _sec_fill = PatternFill(start_color='E8EAF6', end_color='E8EAF6', fill_type='solid')
    def _section(row, title):
        ws.merge_cells(f'A{row}:F{row}')
        c = ws.cell(row=row, column=1, value=title)
        c.font = _sec_font; c.fill = _sec_fill
        for ci in range(1, 7):
            ws.cell(row=row, column=ci).fill = _sec_fill

    # Feature Details
    _section(r, 'SUITE GUIDE')
    r += 1
    _guide_items = [
        ('Summary (this sheet)', 'Overview of the feature, acceptance criteria, coverage breakdown, priority distribution, and data sources.'),
        ('Test Cases', 'All test scenarios with step-by-step actions and expected results. Execute in order — P1 (critical) first.'),
        ('Traceability', 'Maps each Acceptance Criteria to the test cases that cover it. Green = covered, Red = gap.'),
        ('Combinations', 'Device/SIM/Network matrix showing which hardware combos each test case should be run on.'),
    ]
    _cat_guide = [
        ('Happy Path', 'Core positive scenarios — the feature works as designed with valid inputs and expected conditions.'),
        ('Negative', 'Failure and error scenarios — invalid inputs, system failures, timeouts, rollbacks. Verifies graceful handling.'),
        ('Edge Case', 'Unusual but valid scenarios — boundary values, concurrent operations, rare device combos.'),
        ('E2E', 'Full workflow from UI through API to all downstream systems. Validates the complete chain.'),
        ('Regression', 'Ensures existing functionality is not broken by the new feature. Run after every deployment.'),
    ]
    _pri_guide = [
        ('P1 (Critical)', 'Must-run. Core happy paths, rollback, data integrity, E2E flows. Block release if failing.'),
        ('P2 (Important)', 'Should-run. Negative cases, input validation, error handling. High risk if skipped.'),
        ('P3 (Nice-to-have)', 'Good-to-run. UI checks, notifications, low-risk edge cases. Run if time permits.'),
    ]
    ws.cell(row=r, column=1, value='Sheets:').font = _bf
    r += 1
    for sn, desc in _guide_items:
        ws.cell(row=r, column=1, value=sn).font = _bf; ws.cell(row=r, column=1).border = _bdr
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        ws.cell(row=r, column=2, value=desc).font = _nf; ws.cell(row=r, column=2).alignment = _wrap; ws.cell(row=r, column=2).border = _bdr
        r += 1
    r += 1
    ws.cell(row=r, column=1, value='Categories:').font = _bf
    r += 1
    for cn, desc in _cat_guide:
        _cc = CAT_COLORS.get(cn, 'FFFFFF')
        ws.cell(row=r, column=1, value=cn).font = _bf; ws.cell(row=r, column=1).border = _bdr
        ws.cell(row=r, column=1).fill = PatternFill(start_color=_cc, end_color=_cc, fill_type='solid')
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        ws.cell(row=r, column=2, value=desc).font = _nf; ws.cell(row=r, column=2).alignment = _wrap; ws.cell(row=r, column=2).border = _bdr
        r += 1
    r += 1
    ws.cell(row=r, column=1, value='Priorities:').font = _bf
    r += 1
    _pcg = {'P1 (Critical)': 'FFC7CE', 'P2 (Important)': 'FFEB9C', 'P3 (Nice-to-have)': 'C6EFCE'}
    for pn, desc in _pri_guide:
        _pc = _pcg.get(pn, 'FFFFFF')
        ws.cell(row=r, column=1, value=pn).font = _bf; ws.cell(row=r, column=1).border = _bdr
        ws.cell(row=r, column=1).fill = PatternFill(start_color=_pc, end_color=_pc, fill_type='solid')
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        ws.cell(row=r, column=2, value=desc).font = _nf; ws.cell(row=r, column=2).alignment = _wrap; ws.cell(row=r, column=2).border = _bdr
        r += 1
    r += 2

    _section(r, 'FEATURE DETAILS')
    r += 1

    info = [
        ('Feature ID', suite.feature_id),
        ('Summary', suite.feature_title),
        ('Issue Type', 'Epic'),
        ('Priority', suite.jira_priority),
        ('Status', suite.jira_status),
        ('Assignee', suite.jira_assignee),
        ('Reporter', suite.jira_reporter),
        ('Labels', ', '.join(suite.jira_labels)),
        ('Linked Issues', '; '.join(f"{l['key']} - {l['summary'][:80]}" for l in suite.jira_links) if suite.jira_links else 'None'),
        ('Attachments', ', '.join(suite.attachment_names) if suite.attachment_names else 'None'),
        ('PI', suite.pi or 'N/A'),
        ('Channel', ', '.join(suite.channel) if isinstance(suite.channel, list) else suite.channel),
        ('Devices', ', '.join(suite.devices)),
        ('Networks', ', '.join(suite.networks)),
    ]
    for label, val in info:
        ws.cell(row=r, column=1, value=label).font = _bf
        ws.cell(row=r, column=1).border = _bdr
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        ws.cell(row=r, column=2, value=val).font = _nf
        ws.cell(row=r, column=2).alignment = _wrap
        ws.cell(row=r, column=2).border = _bdr
        r += 1

    r += 1

    # Scope & Rules
    if suite.scope or suite.rules:
        ws.merge_cells(f'A{r}:F{r}')
        ws.cell(row=r, column=1, value='SCOPE & RULES').font = _bf
        r += 1
        for text in [suite.scope, suite.rules]:
            if text and text.strip():
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
                ws.cell(row=r, column=1, value=text.strip()).font = _nf
                ws.cell(row=r, column=1).alignment = _wrap
                r += 1
        r += 1

    # Acceptance Criteria
    if suite.acceptance_criteria:
        ws.merge_cells(f'A{r}:F{r}')
        ws.cell(row=r, column=1, value='ACCEPTANCE CRITERIA (from Jira)').font = _bf
        r += 1
        for i, ac in enumerate(suite.acceptance_criteria, 1):
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
            ws.cell(row=r, column=1, value=f'{i}. {ac}').font = _nf
            ws.cell(row=r, column=1).alignment = _wrap
            r += 1
        r += 1

    # Open Items
    if suite.open_items:
        ws.merge_cells(f'A{r}:F{r}')
        ws.cell(row=r, column=1, value='OPEN ITEMS (from attachments/docs)').font = _bf
        r += 1
        for i, item in enumerate(suite.open_items, 1):
            coverage = suite.open_item_coverage.get(item[:80], 'Not covered')
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
            ws.cell(row=r, column=1, value=f'{i}. {item} → {coverage}').font = _nf
            ws.cell(row=r, column=1).alignment = _wrap
            r += 1
        r += 1

    # Test Case Summary Table — HIGH-LEVEL by group (not individual TCs)
    ws.merge_cells(f'A{r}:F{r}')
    ws.cell(row=r, column=1, value='TEST CASE SUMMARY').font = _bf
    r += 1

    # If groups exist, show group-level summary
    if hasattr(suite, 'groups') and suite.groups and len(suite.groups) > 1:
        for ci, h in enumerate(['Sheet / Group', 'TCs', 'Steps', 'Categories', '', ''], 1):
            c = ws.cell(row=r, column=ci, value=h)
            c.font = _hf; c.fill = _hfill; c.alignment = _center; c.border = _bdr
        r += 1

        total_tcs = 0
        total_steps = 0
        for gname, gtcs in suite.groups.items():
            g_steps = sum(len(tc.steps) for tc in gtcs)
            g_cats = set(tc.category for tc in gtcs)
            ws.cell(row=r, column=1, value=gname).font = _nf
            ws.cell(row=r, column=1).alignment = _wrap; ws.cell(row=r, column=1).border = _bdr
            ws.cell(row=r, column=2, value=str(len(gtcs))).font = _nf
            ws.cell(row=r, column=2).alignment = _center; ws.cell(row=r, column=2).border = _bdr
            ws.cell(row=r, column=3, value=str(g_steps)).font = _nf
            ws.cell(row=r, column=3).alignment = _center; ws.cell(row=r, column=3).border = _bdr
            ws.cell(row=r, column=4, value=', '.join(sorted(g_cats))).font = _nf
            ws.cell(row=r, column=4).alignment = _center; ws.cell(row=r, column=4).border = _bdr
            total_tcs += len(gtcs)
            total_steps += g_steps
            r += 1

        # Totals row — use actual suite count, not group sum (groups may merge/miss)
        _actual_tcs = len(suite.test_cases)
        _actual_steps = sum(len(tc.steps) for tc in suite.test_cases)
        ws.cell(row=r, column=1, value='TOTAL').font = _bf
        ws.cell(row=r, column=1).alignment = _center; ws.cell(row=r, column=1).border = _bdr
        ws.cell(row=r, column=2, value=str(_actual_tcs)).font = _bf
        ws.cell(row=r, column=2).alignment = _center; ws.cell(row=r, column=2).border = _bdr
        ws.cell(row=r, column=3, value=str(_actual_steps)).font = _bf
        ws.cell(row=r, column=3).alignment = _center; ws.cell(row=r, column=3).border = _bdr
        r += 1
    else:
        # Single sheet — show compact summary
        for ci, h in enumerate(['TC#', 'Test Scenario', 'Category', 'Steps', '', ''], 1):
            c = ws.cell(row=r, column=ci, value=h)
            c.font = _hf; c.fill = _hfill; c.alignment = _center; c.border = _bdr
        r += 1
        total_steps = 0
        for tc in suite.test_cases:
            import re as _re3
            short_name = _re3.sub(r'^TC\d+[_\s-]+' + _re3.escape(suite.feature_id) + r'[_\s-]*', '', tc.summary)
            cat_color = CAT_COLORS.get(tc.category, 'FFFFFF')
            ws.cell(row=r, column=1, value=f'TC{tc.sno.zfill(2)}').font = _nf
            ws.cell(row=r, column=1).alignment = _center; ws.cell(row=r, column=1).border = _bdr
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
            ws.cell(row=r, column=2, value=short_name[:80]).font = _nf
            ws.cell(row=r, column=2).alignment = _wrap; ws.cell(row=r, column=2).border = _bdr
            ws.cell(row=r, column=4, value=tc.category).font = _nf
            ws.cell(row=r, column=4).alignment = _center; ws.cell(row=r, column=4).border = _bdr
            ws.cell(row=r, column=4).fill = PatternFill(start_color=cat_color, end_color=cat_color, fill_type='solid')
            ws.cell(row=r, column=5, value=str(len(tc.steps))).font = _nf
            ws.cell(row=r, column=5).alignment = _center; ws.cell(row=r, column=5).border = _bdr
            total_steps += len(tc.steps)
            r += 1
        ws.cell(row=r, column=1, value='TOTAL').font = _bf
        ws.cell(row=r, column=1).alignment = _center; ws.cell(row=r, column=1).border = _bdr
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        ws.cell(row=r, column=2, value=f'{len(suite.test_cases)} Test Cases | {total_steps} Test Steps').font = _bf
        ws.cell(row=r, column=2).alignment = _center; ws.cell(row=r, column=2).border = _bdr
        r += 1
    r += 2

    # Coverage Breakdown
    ws.merge_cells(f'A{r}:F{r}')
    ws.cell(row=r, column=1, value='COVERAGE BREAKDOWN').font = _bf
    r += 1
    cats = {}
    for tc in suite.test_cases:
        cats.setdefault(tc.category, []).append(f'TC{tc.sno.zfill(2)}')
    for ci, h in enumerate(['Category', 'Count', 'Test Cases'], 1):
        c = ws.cell(row=r, column=ci, value=h)
        c.font = _hf; c.fill = _hfill; c.alignment = _center; c.border = _bdr
    r += 1
    for cat, tcs in cats.items():
        ws.cell(row=r, column=1, value=cat).font = _nf
        ws.cell(row=r, column=1).alignment = _center; ws.cell(row=r, column=1).border = _bdr
        cat_color = CAT_COLORS.get(cat, 'FFFFFF')
        ws.cell(row=r, column=1).fill = PatternFill(start_color=cat_color, end_color=cat_color, fill_type='solid')
        ws.cell(row=r, column=2, value=str(len(tcs))).font = _nf
        ws.cell(row=r, column=2).alignment = _center; ws.cell(row=r, column=2).border = _bdr
        ws.cell(row=r, column=3, value=', '.join(tcs)).font = _nf
        ws.cell(row=r, column=3).alignment = _center; ws.cell(row=r, column=3).border = _bdr
        r += 1

    r += 1

    # Priority Distribution (V4)
    pris = {}
    for tc in suite.test_cases:
        pri = getattr(tc, '_priority', 'P3')
        pris.setdefault(pri, []).append(f'TC{tc.sno.zfill(2)}')
    if pris:
        ws.merge_cells(f'A{r}:F{r}')
        ws.cell(row=r, column=1, value='PRIORITY DISTRIBUTION').font = _bf
        r += 1
        _pri_colors = {'P1': 'FFC7CE', 'P2': 'FFEB9C', 'P3': 'C6EFCE'}
        for ci, h in enumerate(['Priority', 'Count', 'Test Cases'], 1):
            c = ws.cell(row=r, column=ci, value=h)
            c.font = _hf; c.fill = _hfill; c.alignment = _center; c.border = _bdr
        r += 1
        for pri in ['P1', 'P2', 'P3']:
            if pri in pris:
                ws.cell(row=r, column=1, value=pri).font = _nf
                ws.cell(row=r, column=1).alignment = _center; ws.cell(row=r, column=1).border = _bdr
                _pc = _pri_colors.get(pri, 'FFFFFF')
                ws.cell(row=r, column=1).fill = PatternFill(start_color=_pc, end_color=_pc, fill_type='solid')
                ws.cell(row=r, column=2, value=str(len(pris[pri]))).font = _nf
                ws.cell(row=r, column=2).alignment = _center; ws.cell(row=r, column=2).border = _bdr
                ws.cell(row=r, column=3, value=', '.join(pris[pri][:20])).font = _nf
                ws.cell(row=r, column=3).alignment = _center; ws.cell(row=r, column=3).border = _bdr
                r += 1
        r += 1

    # Warnings
    if suite.warnings:
        ws.merge_cells(f'A{r}:F{r}')
        ws.cell(row=r, column=1, value='⚠️ WARNINGS').font = _bf
        r += 1
        for w in suite.warnings:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
            ws.cell(row=r, column=1, value=w).font = _nf
            ws.cell(row=r, column=1).alignment = _wrap
            r += 1
        r += 1

    # Data Sources
    ws.merge_cells(f'A{r}:F{r}')
    ws.cell(row=r, column=1, value='DATA SOURCES').font = _bf
    r += 1
    for src in suite.data_sources:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        ws.cell(row=r, column=1, value=src).font = _nf
        ws.cell(row=r, column=1).alignment = _wrap
        r += 1


def _build_testcases_sheet(wb, suite: TestSuite, sheet_name=None, tc_subset=None):
    """Build a Test Cases sheet. If tc_subset provided, only those TCs are included.
    TCs are renumbered per-sheet starting from 1."""
    ws = wb.create_sheet(sheet_name or suite.feature_id)
    tcs = tc_subset if tc_subset else suite.test_cases
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 48
    ws.column_dimensions['C'].width = 55
    ws.column_dimensions['D'].width = 48
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 60
    ws.column_dimensions['G'].width = 55
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 18
    ws.column_dimensions['J'].width = 18

    # Row 1: Headers — bold white on soft navy (no feature description banner)
    ws.append(EXCEL_HEADERS)
    _header_fill = PatternFill(start_color='1A237E', end_color='1A237E', fill_type='solid')  # deep indigo
    _header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    for ci in range(1, len(EXCEL_HEADERS) + 1):
        c = ws.cell(row=1, column=ci)
        c.font = _header_font
        c.fill = _header_fill
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = _bdr

    # Freeze panes — keep Row 1 (headers) visible
    ws.freeze_panes = 'A2'

    # Write test cases (renumbered per-sheet starting from 1)
    tc_start_rows = []
    row = 2
    for sheet_idx, tc in enumerate(tcs, 1):
        tc_start_rows.append(row)
        # Per-sheet S.No
        sheet_sno = str(sheet_idx)
        # Clean summary: replace old global TC number with per-sheet number
        import re as _re2
        clean_summary = _re2.sub(r'^TC\d+_', 'TC%02d_' % sheet_idx, tc.summary)
        for si, step in enumerate(tc.steps):
            if si == 0:
                ws.cell(row=row, column=1, value=sheet_sno).alignment = _wrap
                ws.cell(row=row, column=2, value=clean_summary).alignment = _wrap
                ws.cell(row=row, column=3, value=tc.description).alignment = _wrap
                ws.cell(row=row, column=4, value=tc.preconditions).alignment = _wrap
                ws.cell(row=row, column=8, value=tc.story_linkage).alignment = _wrap
                ws.cell(row=row, column=9, value=tc.label).alignment = _wrap
                ws.cell(row=row, column=10, value=tc.story_linkage).alignment = _wrap
            ws.cell(row=row, column=5, value=step.step_num).alignment = _wrap
            ws.cell(row=row, column=6, value=step.summary).alignment = _wrap
            ws.cell(row=row, column=7, value=step.expected).alignment = _wrap
            for c in range(1, len(EXCEL_HEADERS) + 1):
                ws.cell(row=row, column=c).border = _bdr
            row += 1
    tc_start_rows.append(row)  # sentinel

    # Merge columns per test case
    for i in range(len(tc_start_rows) - 1):
        sr = tc_start_rows[i]
        er = tc_start_rows[i + 1] - 1
        if er <= sr:
            continue
        for col in MERGE_COLS:
            ws.merge_cells(start_row=sr, start_column=col, end_row=er, end_column=col)
            cell = ws.cell(row=sr, column=col)
            if col in (1, 8, 9, 10):
                cell.alignment = _center
            else:
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    # Category-based row coloring + bold S.No with category color
    for i in range(len(tc_start_rows) - 1):
        sr = tc_start_rows[i]
        er = tc_start_rows[i + 1] - 1
        # Get category of this TC
        cat = tcs[i].category if i < len(tcs) else 'Happy Path'
        row_fill = _CAT_ROW_FILLS.get(cat, _lb if i % 2 == 0 else _wf)
        sno_font = _CAT_SNO_FONTS.get(cat, _bf)
        for r in range(sr, er + 1):
            for c in range(1, len(EXCEL_HEADERS) + 1):
                ws.cell(row=r, column=c).fill = row_fill
        # Bold S.No with category color
        ws.cell(row=sr, column=1).font = sno_font
        # Bold summary
        ws.cell(row=sr, column=2).font = Font(name='Calibri', bold=True, size=11)


def _build_traceability_sheet(wb, suite: TestSuite):
    """Build AC → TC traceability matrix."""
    if not suite.ac_traceability:
        return

    ws = wb.create_sheet('Traceability')
    ws.freeze_panes = 'A4'  # Point 16: Freeze header rows
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 70
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 15

    # Title
    ws.merge_cells('A1:D1')
    ws.cell(row=1, column=1, value=f'ACCEPTANCE CRITERIA TRACEABILITY - {suite.feature_id}')
    ws.cell(row=1, column=1).font = Font(name='Calibri', bold=True, size=14, color=NAVY)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')

    # Headers
    for ci, h in enumerate(['AC#', 'Acceptance Criteria', 'Covering Test Cases', 'Status'], 1):
        c = ws.cell(row=3, column=ci, value=h)
        c.font = _hf; c.fill = _hfill; c.alignment = _center; c.border = _bdr

    r = 4
    for i, (ac, tcs) in enumerate(suite.ac_traceability.items(), 1):
        ws.cell(row=r, column=1, value=f'AC{i}').font = _nf
        ws.cell(row=r, column=1).alignment = _center; ws.cell(row=r, column=1).border = _bdr
        ws.cell(row=r, column=2, value=ac).font = _nf
        ws.cell(row=r, column=2).alignment = _wrap; ws.cell(row=r, column=2).border = _bdr
        ws.cell(row=r, column=3, value=', '.join(tcs)).font = _nf
        ws.cell(row=r, column=3).alignment = _center; ws.cell(row=r, column=3).border = _bdr

        has_coverage = not any('NO COVERAGE' in t for t in tcs)
        status = '✅ Covered' if has_coverage else '⚠️ Gap'
        ws.cell(row=r, column=4, value=status).font = _nf
        ws.cell(row=r, column=4).alignment = _center; ws.cell(row=r, column=4).border = _bdr
        if not has_coverage:
            ws.cell(row=r, column=4).fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        else:
            ws.cell(row=r, column=4).fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        r += 1


def _build_combinations_sheet(wb, suite: TestSuite):
    """Build the Combinations sheet listing all device matrix combinations."""
    ws = wb.create_sheet('Combinations')
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 30

    # Headers
    headers = ['Combination_ID', 'Channel', 'Device', 'OS', 'SIM Type', 'Network', 'Key']
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = _hf; c.fill = _hfill; c.alignment = _center; c.border = _bdr

    # Data
    for ri, combo in enumerate(suite.combinations, 2):
        ws.cell(row=ri, column=1, value=combo.get('id', '')).border = _bdr
        ws.cell(row=ri, column=1).alignment = _center
        ws.cell(row=ri, column=2, value=combo.get('channel', '')).border = _bdr
        ws.cell(row=ri, column=2).alignment = _center
        ws.cell(row=ri, column=3, value=combo.get('device', '')).border = _bdr
        ws.cell(row=ri, column=3).alignment = _center
        ws.cell(row=ri, column=4, value=combo.get('os', '')).border = _bdr
        ws.cell(row=ri, column=4).alignment = _center
        ws.cell(row=ri, column=5, value=combo.get('sim', '')).border = _bdr
        ws.cell(row=ri, column=5).alignment = _center
        ws.cell(row=ri, column=6, value=combo.get('network', '')).border = _bdr
        ws.cell(row=ri, column=6).alignment = _center
        ws.cell(row=ri, column=7, value=combo.get('key', '')).border = _bdr
        ws.cell(row=ri, column=7).alignment = _center
