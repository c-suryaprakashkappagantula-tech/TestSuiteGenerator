"""
cabot_excel_generator.py — Generate an Excel file matching the EXACT format
of Cabot_NMP_MOBIT2-62376_TestCases_v3.xlsx.

Columns:
  unique_id | type | name | step_type | step_description | test_type |
  product_areas | covered_content | designer | description |
  estimated_duration | owner | phase | user_tags

Row structure per TC:
  • One row with type="test_manual", name=tc_name, all TC-level fields populated
  • One row per step with type="step", name="simple",
    step_type=step_type, step_description=step_description

Part of TSG V6.0 — Cabot Test Suite Engine rebuild.
"""
from __future__ import annotations

from pathlib import Path
from typing import List
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from .endpoint_tc_generator import CabotTestCase, CabotTestStep


# ================================================================
# COLUMN LAYOUT — matches Cabot_NMP sample exactly
# ================================================================

CABOT_COLUMNS = [
    "unique_id",
    "type",
    "name",
    "step_type",
    "step_description",
    "test_type",
    "product_areas",
    "covered_content",
    "designer",
    "description",
    "estimated_duration",
    "owner",
    "phase",
    "user_tags",
]

# Column widths (approximate, tuned for readability)
_COL_WIDTHS = {
    "A": 12,   # unique_id
    "B": 16,   # type
    "C": 70,   # name
    "D": 12,   # step_type
    "E": 80,   # step_description
    "F": 60,   # test_type
    "G": 28,   # product_areas
    "H": 18,   # covered_content
    "I": 18,   # designer
    "J": 60,   # description
    "K": 20,   # estimated_duration
    "L": 18,   # owner
    "M": 14,   # phase
    "N": 22,   # user_tags
}

# ================================================================
# STYLES
# ================================================================

_HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

_TC_FONT = Font(name="Calibri", bold=True, size=11)
_TC_FILL = PatternFill(start_color="E8EAF6", end_color="E8EAF6", fill_type="solid")

_STEP_FONT = Font(name="Calibri", size=11)
_STEP_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
_STEP_ALT_FILL = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

_WRAP = Alignment(wrap_text=True, vertical="top")
_CENTER = Alignment(horizontal="center", vertical="top", wrap_text=True)

_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


# ================================================================
# EXCEL GENERATION
# ================================================================

def generate_cabot_excel(
    test_cases: List[CabotTestCase],
    feature_id: str,
    output_dir: Path,
    designer: str = "",
    owner: str = "",
    phase: str = "",
) -> Path:
    """Generate an Excel file in the Cabot_NMP format.
    Matches the exact layout of Cabot_NMP_MOBIT2-62376_TestCases_v3.xlsx:
      - Sheet name: "manual tests"
      - unique_id: sequential across ALL rows (TC + step rows)
      - TC rows: type="test_manual", name=tc_name, test_type="API"
      - Step rows: type="step", name=None, step_type/step_description populated
    """
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = "Cabot_NMP_%s_TestCases_%s.xlsx" % (feature_id, ts)
    out_path = output_dir / filename

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "manual tests"

    # ── Column widths ──
    for col_letter, width in _COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # ── Header row ──
    for ci, header in enumerate(CABOT_COLUMNS, 1):
        cell = ws.cell(row=1, column=ci, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _BORDER

    ws.freeze_panes = "A2"

    # ── Data rows — unique_id is sequential across ALL rows ──
    row = 2
    uid = 1
    for tc_idx, tc in enumerate(test_cases):
        # TC-level row: type = "test_manual"
        _write_tc_row(ws, row, tc, uid, designer, owner, phase)
        uid += 1
        row += 1

        # Step rows: type = "step"
        for step_idx, step in enumerate(tc.steps):
            _write_step_row(ws, row, step, uid, tc_idx)
            uid += 1
            row += 1

    wb.save(str(out_path))
    return out_path


# ================================================================
# ROW WRITERS
# ================================================================

def _write_tc_row(
    ws,
    row: int,
    tc: CabotTestCase,
    uid: int,
    designer: str,
    owner: str,
    phase: str,
) -> None:
    """Write a single TC-level row (type=test_manual).
    Matches sample: unique_id, type=test_manual, name=tc_name, test_type=API,
    product_areas=folder_path, description=long sentence, user_tags=tag_chain.
    step_type and step_description are None for TC rows."""
    values = {
        1: uid,                             # unique_id (sequential)
        2: "test_manual",                   # type
        3: tc.tc_name,                      # name
        4: None,                            # step_type (None for TC row per sample)
        5: None,                            # step_description (None for TC row per sample)
        6: tc.test_type,                    # test_type = "API"
        7: tc.product_areas,                # product_areas = full folder path
        8: None,                            # covered_content (None per sample)
        9: designer or None,                # designer
        10: tc.description,                 # description = long validation sentence
        11: tc.estimated_duration,          # estimated_duration
        12: owner or None,                  # owner
        13: phase or None,                  # phase
        14: tc.user_tags,                   # user_tags = full tag chain
    }
    for ci, val in values.items():
        cell = ws.cell(row=row, column=ci, value=val)
        cell.font = _TC_FONT
        cell.fill = _TC_FILL
        cell.border = _BORDER
        if ci in (1, 2, 8, 11):
            cell.alignment = _CENTER
        else:
            cell.alignment = _WRAP


def _write_step_row(ws, row: int, step: CabotTestStep, uid: int, tc_idx: int) -> None:
    """Write a single step row (type=step).
    Matches sample: unique_id=sequential, type=step, name=None,
    step_type=simple/validation, step_description=text.
    All other columns are None."""
    fill = _STEP_FILL if tc_idx % 2 == 0 else _STEP_ALT_FILL
    values = {
        1: uid,                             # unique_id (sequential)
        2: "step",                          # type
        3: None,                            # name (None for step rows per sample)
        4: step.step_type,                  # step_type ("simple" or "validation")
        5: step.step_description,           # step_description
        6: None,                            # test_type (None)
        7: None,                            # product_areas (None)
        8: None,                            # covered_content (None)
        9: None,                            # designer (None)
        10: None,                           # description (None)
        11: None,                           # estimated_duration (None)
        12: None,                           # owner (None)
        13: None,                           # phase (None)
        14: None,                           # user_tags (None)
    }
    for ci, val in values.items():
        cell = ws.cell(row=row, column=ci, value=val)
        cell.font = _STEP_FONT
        cell.fill = fill
        cell.border = _BORDER
        if ci in (1, 2, 3, 4):
            cell.alignment = _CENTER
        else:
            cell.alignment = _WRAP
